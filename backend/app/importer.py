"""MS Forms Excel importer.

Resilient to column reordering by matching headers (case-insensitive substring).
Handles the 2024 RealHack format and the 2026 form structure (which adds email
per member).
"""
from __future__ import annotations

import re
from datetime import datetime
from typing import Any, Iterable

from openpyxl import load_workbook

from .models import Team, Member


def _norm(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _find_col(headers: list[str], *needles: str) -> int | None:
    """Return the first column index whose header contains ALL given substrings."""
    for i, h in enumerate(headers):
        nh = _norm(h)
        if all(_norm(n) in nh for n in needles):
            return i
    return None


def _val(row: tuple, idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def parse_workbook(path: str) -> list[dict]:
    """Parse the first sheet of the MS Forms Excel export into a list of team dicts."""
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "") for h in rows[0]]

    # Submitter/meta
    col_id = _find_col(headers, "id")
    col_completed = _find_col(headers, "completion time")
    col_team = _find_col(headers, "team name")
    # Mentor — match either "team mentor" or "mentor name"
    col_mentor_name = _find_col(headers, "mentor name") or _find_col(headers, "team mentor")
    col_mentor_email = _find_col(headers, "mentor", "email")
    col_mentor_location = _find_col(headers, "mentor", "location")
    col_mentor_tshirt = _find_col(headers, "mentor", "shirt")
    # Mailing address — captured by the MS Forms column
    # "Enter your mailing address if you opted for US or PH as location"
    # (and a mentor-tagged variant when the form includes it).
    col_mentor_address = (
        _find_col(headers, "mentor", "mailing", "address")
        or _find_col(headers, "mentor", "address")
    )

    # Idea/approach fields
    col_idea = _find_col(headers, "idea") or _find_col(headers, "problem statement")
    col_tools = _find_col(headers, "tools") or _find_col(headers, "tech stack")
    col_approach = _find_col(headers, "approach")
    col_viability = _find_col(headers, "viability")
    col_biz_value = _find_col(headers, "business value")

    # Members — detect 1..5 by header patterns. Two known shapes:
    #   2024: "Teammate 1", "Location", "T-Shirt Size ..."  (no member email)
    #   2026: "Member 1 Name", "Member 1 Email", "Member 1 T-shirt Size", "Member 1 Location"
    members_meta: list[dict[str, int | None]] = []
    for n in range(1, 6):
        name_idx = (
            _find_col(headers, f"member {n} name")
            or _find_col(headers, f"member{n} name")
            or _find_col(headers, f"teammate {n}")
        )
        if name_idx is None:
            members_meta.append({"name": None, "email": None, "tshirt": None, "location": None})
            continue

        # Member-specific columns may be tagged with the slot number or sit
        # immediately to the right in 2024-style export. We probe both.
        email_idx = _find_col(headers, f"member {n}", "email") or _find_col(headers, f"member{n}", "email")
        tshirt_idx = (
            _find_col(headers, f"member {n}", "shirt")
            or _find_col(headers, f"member{n}", "shirt")
        )
        location_idx = (
            _find_col(headers, f"member {n}", "location")
            or _find_col(headers, f"member{n}", "location")
        )
        address_idx = (
            _find_col(headers, f"member {n}", "mailing", "address")
            or _find_col(headers, f"member{n}", "mailing", "address")
            or _find_col(headers, f"member {n}", "address")
            or _find_col(headers, f"member{n}", "address")
        )

        # 2024 fallback: columns immediately after the teammate name are
        # Location and T-shirt Size for that teammate.
        if tshirt_idx is None or location_idx is None:
            if location_idx is None:
                # the column right after the name is usually location
                candidate = name_idx + 1
                if candidate < len(headers) and "location" in _norm(headers[candidate]):
                    location_idx = candidate
            if tshirt_idx is None:
                candidate = name_idx + 2
                if candidate < len(headers) and "shirt" in _norm(headers[candidate]):
                    tshirt_idx = candidate

        members_meta.append(
            {
                "name": name_idx,
                "email": email_idx,
                "tshirt": tshirt_idx,
                "location": location_idx,
                "address": address_idx,
            }
        )

    out: list[dict] = []
    for r in rows[1:]:
        if not any(v not in (None, "") for v in r):
            continue  # empty row

        team_name = _val(r, col_team)
        if not team_name:
            continue  # skip rows without a team name

        members: list[dict] = []
        for slot, meta in enumerate(members_meta, start=1):
            name = _val(r, meta["name"])
            if not name:
                continue
            members.append(
                {
                    "name": str(name),
                    "email": _val(r, meta["email"]),
                    "location": _val(r, meta["location"]),
                    "tshirt_size": _val(r, meta["tshirt"]),
                    "address": _val(r, meta["address"]),
                    "position": slot,
                }
            )

        submitted_at = _val(r, col_completed)
        if submitted_at and not isinstance(submitted_at, datetime):
            submitted_at = None

        external_id = _val(r, col_id)

        out.append(
            {
                "external_id": str(external_id) if external_id is not None else None,
                "name": str(team_name),
                "mentor_name": _val(r, col_mentor_name),
                "mentor_email": _val(r, col_mentor_email),
                "mentor_location": _val(r, col_mentor_location),
                "mentor_tshirt_size": _val(r, col_mentor_tshirt),
                "mentor_address": _val(r, col_mentor_address),
                "idea": _val(r, col_idea),
                "tools": _val(r, col_tools),
                "approach": _val(r, col_approach),
                "viability": _val(r, col_viability),
                "business_value": _val(r, col_biz_value),
                "submitted_at": submitted_at,
                "raw": {h: (r[i] if i < len(r) else None) for i, h in enumerate(headers)},
                "members": members,
            }
        )
    return out


def dicts_to_models(team_dicts: Iterable[dict]) -> list[Team]:
    teams: list[Team] = []
    for d in team_dicts:
        team = Team(
            external_id=d.get("external_id"),
            name=d["name"],
            mentor_name=d.get("mentor_name"),
            mentor_email=d.get("mentor_email"),
            mentor_location=d.get("mentor_location"),
            mentor_tshirt_size=d.get("mentor_tshirt_size"),
            mentor_address=d.get("mentor_address"),
            idea=d.get("idea"),
            tools=d.get("tools"),
            approach=d.get("approach"),
            viability=d.get("viability"),
            business_value=d.get("business_value"),
            submitted_at=d.get("submitted_at"),
            raw=_jsonify(d.get("raw")),
        )
        for m in d.get("members", []):
            team.members.append(
                Member(
                    name=m["name"],
                    email=m.get("email"),
                    location=m.get("location"),
                    tshirt_size=m.get("tshirt_size"),
                    address=m.get("address"),
                    position=m.get("position", 0),
                )
            )
        teams.append(team)
    return teams


def _jsonify(d: dict | None) -> dict | None:
    if d is None:
        return None
    out: dict = {}
    for k, v in d.items():
        if isinstance(v, datetime):
            out[k] = v.isoformat()
        else:
            out[k] = v
    return out
