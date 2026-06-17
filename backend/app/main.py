from __future__ import annotations

import logging
import os
import shutil
import tempfile
import threading
import uuid
from collections import Counter
from datetime import datetime
from typing import Any

import csv
import io

from fastapi import FastAPI, File, HTTPException, Request, UploadFile, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy import text, func
from sqlalchemy.orm import Session

from .config import settings
from .db import Base, SessionLocal, engine, get_db, lightweight_migrate
from . import models
from .importer import parse_workbook, dicts_to_models
from .screener import screen_all
from .ai_screener import score_all as ai_score_all, score_team as ai_score_team
from .emails import TEMPLATES as EMAIL_TEMPLATES, render_many as render_emails
from .judge import ai_judge, merge_human_scores
from . import judging
from . import comms
from . import backup as backup_service
from . import llm as llm_service
from . import chat as chat_service
from . import scheduling
from .auth import require_auth, fetch_profile, build_profile_payload
from fastapi import Security
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer

_bearer_scheme = HTTPBearer(auto_error=False)
from .schemas import (
    TeamOut, UploadResult, DashboardStats, AIScreenResult,
    EmailTemplateOut, EmailRenderRequest, RenderedEmail,
    JudgeAIRequest, JudgeHumanRequest,
    JudgeOut, JudgeCreate, JudgeUpdate, JudgeScoreSubmit, JudgeScoreOut,
    JudgeBulkRequest, JudgeBulkResult,
    TeamSeatRequest,
    SwagExtraOut, SwagExtraImportResult,
    UserRoleOut, JudgeAssignmentSet, JudgeAssignmentOut,
    PanelOut, PanelCreate, PanelUpdate, PanelTeamsSet, PanelJudgesSet, PanelSwapTeamDays,
    AdoptChannelByLinkRequest,
    SwagPersonOut, SwagMarkRequest, SwagStats,
    LeaderboardOut, LeaderboardRow, JUDGE_RUBRIC_AXES,
    CommLogOut, TeamChannelCreateRequest, TeamMessageRequest, BroadcastRequest,
    CommLogCreateRequest, RepoCheckOut, ReadinessFlagsRequest,
    ChatRequest, ChatResponse,
    RoundAdvanceRequest, WinnersSetRequest, RoundSummary,
    TeamCreate, TeamPatch, MemberCreate, MemberPatch, MemberOut,
)


app = FastAPI(title="RealHack Pilot API", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.cors_origins_list,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---- Auth middleware ----
# Every /api/* request requires a valid Entra Bearer token. Health + docs are
# exempt so monitoring + dev-time exploration still work without sign-in.
_AUTH_EXEMPT_PATHS = {"/api/health", "/docs", "/redoc", "/openapi.json"}

# Paths where ANY signed-in RealPage AAD user is allowed, regardless of
# whether they're an organizer or judge in our DB. Used for the swag-staff
# flow: volunteers who sign in with their normal AAD account but aren't on
# the dashboard roster get a stripped-down 'mark collected' UI scoped to
# these endpoints. Everything NOT in this set requires the user to be in the
# `judges` table (any role) or be a sandbox admin -- enforced below.
#
# This used to be guaranteed by Entra's 'Assignment Required = Yes' on the
# Enterprise App. Once that's flipped to No (so swag staff can sign in), the
# token alone no longer proves the user belongs here -- the role check
# below replaces that guarantee at the application layer.
_SIGNED_IN_OK_PATHS = {
    "/api/me",
    "/api/me/role",
    "/api/swag/people",
    "/api/swag/mark",
    "/api/sandbox/mode",
}


def _lookup_role_for_email(email: str) -> str | None:
    """Returns the role string from the judges table for this email, or
    'organizer' for sandbox admins, or None if unknown. Used by the
    middleware to gate non-allowlist endpoints. Opens its own DB session
    because middleware runs outside of Depends().
    """
    email = (email or "").strip().lower()
    if not email:
        return None
    if email in SANDBOX_ADMIN_EMAILS:
        return "organizer"
    from .db import SessionLocal
    with SessionLocal() as db:
        row = (
            db.query(models.Judge)
            .filter(func.lower(models.Judge.email) == email)
            .filter(models.Judge.is_active.is_(True))
            .first()
        )
        return (row.role or "judge") if row else None


# Roles that can read general dashboard data + perform actions on team /
# scoring / comms endpoints. REWS is intentionally excluded -- they have
# the AAD group membership to sign in but are scoped to pickup-desk swag
# operations only.
_DASHBOARD_ROLES = {"organizer", "judge"}

# Roles allowed to mark swag collected (no Undo). Organizers + judges can
# also do this (they're typically the ones at the desk if no REWS is there).
_SWAG_DESK_ROLES = {"organizer", "judge", "rews"}


@app.middleware("http")
async def _require_auth_middleware(request, call_next):
    path = request.url.path
    # Allow CORS preflight and exempt paths through without auth.
    # /api/public/* is open by design — it powers the QR-code judging-walk
    # page where a judge scans a printed code with no login required.
    if (
        request.method == "OPTIONS"
        or path in _AUTH_EXEMPT_PATHS
        or path.startswith("/api/public/")
        or not path.startswith("/api/")
    ):
        return await call_next(request)
    auth_header = request.headers.get("Authorization") or ""
    if not auth_header.lower().startswith("bearer "):
        return JSONResponse(
            status_code=401,
            content={"detail": "Authentication required"},
            headers={"WWW-Authenticate": "Bearer"},
        )
    token = auth_header.split(" ", 1)[1].strip()
    try:
        from .auth import validate_token  # local import to avoid circular
        claims = validate_token(token)
    except Exception as e:
        return JSONResponse(
            status_code=401,
            content={"detail": f"Invalid token: {e}"},
            headers={"WWW-Authenticate": "Bearer"},
        )

    # Role gate, three tiers:
    #   1. _SIGNED_IN_OK_PATHS -- any valid AAD token works. Used so the
    #      'Not registered' card has somewhere to call (/api/me/role) and
    #      so the frontend can negotiate its routing decision.
    #   2. /api/swag/{people,mark} -- organizer OR judge OR rews. REWS
    #      volunteers can mark people collected. (unmark gates organizer-
    #      only inside the handler itself.)
    #   3. Everything else under /api/* -- organizer OR judge. Excludes
    #      rews so a volunteer who somehow crafts a curl can't read team
    #      data or fire comms.
    if path not in _SIGNED_IN_OK_PATHS:
        email = (
            claims.get("preferred_username")
            or claims.get("upn")
            or claims.get("email")
            or ""
        )
        role = _lookup_role_for_email(email)
        allowed = _SWAG_DESK_ROLES if path.startswith("/api/swag/") else _DASHBOARD_ROLES
        if role not in allowed:
            return JSONResponse(
                status_code=403,
                content={
                    "detail": (
                        "You're signed in but don't have the role required for this action. "
                        "If you're helping with swag pickup, an organizer needs to add you "
                        "to the Judges & Organizers panel with role 'REWS'."
                    ),
                },
            )

    return await call_next(request)


@app.on_event("startup")
def _startup() -> None:
    # Wrap create_all in try/except: if the DB user lacks CREATE on the
    # public schema (same trap as the ALTER permission issue -- the
    # realhack_pilot_app user only has DML grants), we don't want the
    # app to crash at startup. The relevant table simply won't exist
    # and the endpoints that touch it will 500 individually; everything
    # else keeps working. Organizer fixes ownership out-of-band via psql.
    _startup_log = logging.getLogger("startup")
    try:
        Base.metadata.create_all(bind=engine)
    except Exception as e:
        _startup_log.error("create_all on prod DB failed (likely missing CREATE on schema): %s", e)
    lightweight_migrate()
    # Mirror the schema onto the sandbox DB if one is configured. Sandbox content
    # itself is populated on demand via POST /api/admin/sandbox/refresh.
    from .db import sandbox_engine
    if sandbox_engine is not None:
        try:
            Base.metadata.create_all(bind=sandbox_engine)
        except Exception as e:
            _startup_log.error("create_all on sandbox DB failed: %s", e)
        # Run the same lightweight column migrations against sandbox so it gets
        # any columns added to the model after the sandbox was first created.
        lightweight_migrate(target_engine=sandbox_engine)
    backup_service.start_scheduler()


@app.get("/api/health")
def health() -> dict:
    return {"status": "ok", "env": settings.app_env}


# ===== Sandbox / Test Mode =====
#
# Test Mode is a super-admin-only toggle on the frontend. When on, every API
# call from that browser sends `x-sandbox: true`; the `get_db` dependency
# routes the session to the sandbox database. Production data is never read
# or written from a Test-Mode session. The refresh endpoint copies the live
# prod rows into the sandbox so the organizer can try destructive flows
# (advance teams, crown finalists, send mocked emails) against realistic
# data without affecting the live event.

# Lowercased list — same set used to gate the 'Preview as judge' control.
SANDBOX_ADMIN_EMAILS = {"shaikshavali.kalluri@realpage.com"}


def _require_sandbox_admin(claims: dict, creds: HTTPAuthorizationCredentials) -> str:
    """Raise unless the caller is a super-admin allowed to manage the sandbox."""
    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    email = (profile.get("email") or "").strip().lower()
    if email not in SANDBOX_ADMIN_EMAILS:
        raise HTTPException(status_code=403, detail="Sandbox controls are super-admin only")
    return email


@app.get("/api/admin/sandbox/status")
def sandbox_status(
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
) -> dict:
    """Report whether sandbox is configured + (if so) basic row counts."""
    _require_sandbox_admin(claims, creds)
    from .db import sandbox_engine, SandboxSessionLocal
    if sandbox_engine is None or SandboxSessionLocal is None:
        return {"configured": False, "message": "SANDBOX_DATABASE_URL not set on the server."}
    db = SandboxSessionLocal()
    try:
        counts = {
            "teams": db.query(models.Team).count(),
            "members": db.query(models.Member).count(),
            "judges": db.query(models.Judge).count(),
            "panels": db.query(models.Panel).count(),
            "judge_scores": db.query(models.JudgeScore).count(),
        }
    finally:
        db.close()
    return {"configured": True, "counts": counts}


@app.post("/api/admin/sandbox/refresh")
def sandbox_refresh(
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
) -> dict:
    """Wipe the sandbox database and reload it with a fresh copy of prod data.

    Streams rows table-by-table via SQLAlchemy reflection — no shell tooling
    needed, works against the same Postgres server. Tables are wiped in FK-safe
    order, then refilled in dependency order.
    """
    _require_sandbox_admin(claims, creds)
    from .db import sandbox_engine, SandboxSessionLocal
    if sandbox_engine is None or SandboxSessionLocal is None:
        raise HTTPException(status_code=400, detail="SANDBOX_DATABASE_URL not set on the server")

    # Order matters for FK constraints: wipe children first, refill parents first.
    table_order = [
        models.Team, models.Member, models.Judge, models.Panel,
        models.PanelTeam, models.PanelJudge, models.JudgeAssignment,
        models.JudgeScore, models.CommLog,
    ]

    prod = SessionLocal()
    sb = SandboxSessionLocal()
    copied: dict[str, int] = {}
    try:
        # Wipe sandbox in reverse order so FKs don't trip
        for model in reversed(table_order):
            sb.query(model).delete(synchronize_session=False)
        sb.commit()

        # Copy prod rows over, preserving primary keys
        from sqlalchemy import inspect as sqla_inspect
        for model in table_order:
            rows = prod.query(model).all()
            mapper = sqla_inspect(model)
            col_names = [c.key for c in mapper.columns]
            payload = [{k: getattr(r, k) for k in col_names} for r in rows]
            if payload:
                sb.bulk_insert_mappings(model, payload)
            copied[model.__tablename__] = len(payload)
        sb.commit()
    except Exception as e:
        sb.rollback()
        raise HTTPException(status_code=500, detail=f"Sandbox refresh failed: {e}")
    finally:
        prod.close()
        sb.close()

    return {"refreshed": True, "rows_copied": copied}


@app.get("/api/me", response_model=dict)
def me(
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
) -> dict:
    """Return the signed-in user's profile: name, email, jobTitle, department.

    Combines JWT claims (authoritative for identity) with Graph /me data
    (jobTitle, department) so the dashboard can show a rich profile badge.
    """
    graph_profile = fetch_profile(creds.credentials) if creds else {}
    return build_profile_payload(claims, graph_profile)


@app.get("/api/llm/health")
def llm_health() -> dict:
    return llm_service.health_check()


@app.get("/api/teams", response_model=list[TeamOut])
def list_teams(db: Session = Depends(get_db)) -> list[models.Team]:
    return db.query(models.Team).order_by(models.Team.name.asc()).all()


# ===== Public, no-auth endpoints for the QR-code judging walk =====
#
# These power the printed-card workflow: each team has a QR pointing to
# /team/<id> in the SPA; the page reads from /api/public/teams/<id>
# without requiring login. We surface only the judge-relevant subset and
# strip emails / internal flags / completeness scores. The middleware
# whitelists /api/public/* (see _require_auth_middleware).

def _public_team_dict(team: models.Team, include_idea_full: bool = True) -> dict:
    """Slim, judge-facing view of a team. Excludes emails, screening
    flags, completeness scores, mailing addresses, and other internal data."""
    ai = team.ai_scores or {}
    overall = ai.get("overall") or {}
    # ai_scores["overall"] is {"score": <1-5>, "headline": "..."} -- not a
    # raw number. Extract the score / headline rather than passing the whole
    # object through to the frontend (which would interpolate as
    # '[object Object]/5 OVERALL' in the section heading).
    if isinstance(overall, dict):
        ai_overall_score = overall.get("score")
        ai_overall_headline = overall.get("headline")
    else:
        ai_overall_score = overall  # legacy shape: plain number
        ai_overall_headline = None

    members = [
        {"name": m.name, "location": m.location}
        for m in team.members
        if m.name
    ]
    return {
        "id": team.id,
        "name": team.name,
        "mentor_name": team.mentor_name,
        "idea": team.idea if include_idea_full else (team.idea or "")[:200],
        "tools": team.tools,
        "approach": team.approach,
        "viability": team.viability,
        "business_value": team.business_value,
        "members": members,
        "ai_summary": ai.get("summary"),
        "ai_overall_score": ai_overall_score,
        "ai_overall_headline": ai_overall_headline,
        "seat_floor": team.seat_floor,
        "seat_desk": team.seat_desk,
        "seat_landmark": team.seat_landmark,
        "seat_updated_at": team.seat_updated_at.isoformat() if team.seat_updated_at else None,
        "seat_updated_by": team.seat_updated_by,
    }


@app.get("/api/public/teams")
def public_list_teams(db: Session = Depends(get_db)) -> list[dict]:
    """Lightweight list of every team (id + name + mentor + short idea).
    Powers the bulk 'Print judging cards' page."""
    teams = db.query(models.Team).order_by(models.Team.name.asc()).all()
    return [
        {
            "id": t.id,
            "name": t.name,
            "mentor_name": t.mentor_name,
            "idea_short": (t.idea or "")[:140],
        }
        for t in teams
    ]


@app.get("/api/public/teams/{team_id}")
def public_team_detail(team_id: int, db: Session = Depends(get_db)) -> dict:
    """Judge-facing team detail — read by the QR-code landing page.
    No auth, no emails, no internal flags."""
    team = db.query(models.Team).get(team_id)
    if not team:
        raise HTTPException(status_code=404, detail="team not found")
    return _public_team_dict(team)


VALID_SEAT_FLOORS = {"5th", "9th", "10th"}


@app.post("/api/public/teams/{team_id}/seat")
def public_update_seat(team_id: int, req: TeamSeatRequest, request: Request, db: Session = Depends(get_db)) -> dict:
    """Self-service seat update from the public team page. No auth required:
    anyone with the QR can submit (low blast radius — three short fields, easy
    to fix manually if vandalized). Captures submitter email from the optional
    `x-user-email` header that the SPA sends if MSAL is signed in, for the
    audit trail.
    """
    team = db.query(models.Team).get(team_id)
    if not team:
        raise HTTPException(status_code=404, detail="team not found")

    floor = (req.floor or "").strip()
    if floor not in VALID_SEAT_FLOORS:
        raise HTTPException(status_code=400, detail=f"floor must be one of {sorted(VALID_SEAT_FLOORS)}")
    desk = (req.desk or "").strip()
    if not desk:
        raise HTTPException(status_code=400, detail="desk is required")

    team.seat_floor = floor
    team.seat_desk = desk[:64]
    team.seat_landmark = (req.landmark or "").strip()[:500] or None
    team.seat_updated_at = datetime.utcnow()
    # Submitter attribution: prefer the explicit body field (filled from the
    # form dropdown of mentor + members + 'Someone else'), fall back to the
    # MSAL email header in case the submitter happens to be signed in. Either
    # way it lands in seat_updated_by so organizers can chase the right
    # person if the info turns out to be wrong.
    submitted = (req.submitted_by or "").strip()[:255]
    if not submitted:
        submitted = (request.headers.get("x-user-email") or "").strip().lower()[:255]
    team.seat_updated_by = submitted or None

    db.commit()
    return {
        "id": team.id,
        "seat_floor": team.seat_floor,
        "seat_desk": team.seat_desk,
        "seat_landmark": team.seat_landmark,
        "seat_updated_at": team.seat_updated_at.isoformat() if team.seat_updated_at else None,
        "seat_updated_by": team.seat_updated_by,
    }


@app.get("/api/seat-coverage", response_model=dict)
def seat_coverage(claims: dict = Depends(require_auth), db: Session = Depends(get_db)) -> dict:
    """Floor-walk coverage report for the dashboard: which teams have
    submitted seat info, which haven't. Pending list is alphabetical so
    organizers can hand it to a follow-up chat / Teams nudge."""
    teams = db.query(models.Team).order_by(models.Team.name.asc()).all()
    submitted: list[dict] = []
    pending: list[dict] = []
    by_floor: dict[str, int] = {}
    for t in teams:
        if (t.seat_floor or "").strip() and (t.seat_desk or "").strip():
            submitted.append({
                "id": t.id,
                "name": t.name,
                "floor": t.seat_floor,
                "desk": t.seat_desk,
                "landmark": t.seat_landmark,
                "updated_at": t.seat_updated_at.isoformat() if t.seat_updated_at else None,
                "updated_by": t.seat_updated_by,
            })
            by_floor[t.seat_floor] = by_floor.get(t.seat_floor, 0) + 1
        else:
            pending.append({
                "id": t.id,
                "name": t.name,
                "mentor_name": t.mentor_name,
                "has_channel": bool(t.has_teams_channel and t.teams_channel_id),
            })
    return {
        "total": len(teams),
        "submitted_count": len(submitted),
        "pending_count": len(pending),
        "by_floor": by_floor,
        "submitted": submitted,
        "pending": pending,
    }


@app.get("/api/stats", response_model=DashboardStats)
def stats(db: Session = Depends(get_db)) -> DashboardStats:
    teams = db.query(models.Team).all()
    total = len(teams)
    complete = sum(1 for t in teams if t.completeness_score >= 0.8 and not (t.flags or []))
    flagged = sum(1 for t in teams if t.flags)

    locations: Counter[str] = Counter()
    sizes: Counter[str] = Counter()
    # Deduped people: lowercased email if available, else lowercased name.
    # This is what the dashboard's "Total unique people" tile reports — a
    # person who appears as both a mentor and a team member only counts once.
    unique_people: set[str] = set()
    for t in teams:
        if t.mentor_email:
            unique_people.add(t.mentor_email.strip().lower())
        elif t.mentor_name:
            unique_people.add(t.mentor_name.strip().lower())
        for m in t.members:
            if m.email:
                unique_people.add(m.email.strip().lower())
            elif m.name:
                unique_people.add(m.name.strip().lower())
            if m.location:
                locations[str(m.location).strip()] += 1
            if m.tshirt_size:
                sizes[str(m.tshirt_size).strip().upper()] += 1

    # Cross-team flags via DB scan
    dup_count = 0
    multi_mentor_count = 0
    for t in teams:
        for f in t.flags or []:
            if f.startswith("duplicate_participant:"):
                dup_count += 1
            if f.startswith("mentor_overloaded:"):
                multi_mentor_count += 1

    return DashboardStats(
        total_teams=total,
        complete_teams=complete,
        flagged_teams=flagged,
        duplicate_participants=dup_count,
        multi_team_mentors=multi_mentor_count,
        locations=dict(locations),
        tshirt_sizes=dict(sizes),
        total_unique_people=len(unique_people),
    )


@app.post("/api/upload", response_model=UploadResult)
async def upload_registrations(
    file: UploadFile = File(...),
    replace: bool = True,
    db: Session = Depends(get_db),
) -> UploadResult:
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Expected an .xlsx/.xls file")

    suffix = os.path.splitext(file.filename)[1] or ".xlsx"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        shutil.copyfileobj(file.file, tmp)
        tmp.close()

        team_dicts = parse_workbook(tmp.name)
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass

    skipped = 0
    if replace:
        # Preserve AI scores and organizer-set fields keyed by external_id so
        # they survive the re-upload (TRUNCATE wipes the table but the scores
        # from a previous AI screening run should not be lost).
        preserved: dict[str, dict] = {}
        for existing in db.query(models.Team).all():
            if existing.external_id:
                preserved[existing.external_id] = {
                    "ai_scores": existing.ai_scores,
                    "judge_scores": existing.judge_scores,
                    "advanced_to_round": existing.advanced_to_round,
                    "final_position": existing.final_position,
                    "presentation_uploaded": existing.presentation_uploaded,
                    "repo_url": existing.repo_url,
                    "repo_ready": existing.repo_ready,
                    "repo_check_notes": existing.repo_check_notes,
                    "has_teams_channel": existing.has_teams_channel,
                    "teams_channel_id": existing.teams_channel_id,
                    "teams_channel_created_at": existing.teams_channel_created_at,
                }
        db.execute(text("TRUNCATE TABLE members, teams RESTART IDENTITY CASCADE"))
        db.flush()

    teams = dicts_to_models(team_dicts)
    # Deduplicate on external_id — MS Forms exports occasionally contain duplicate rows.
    seen_ext_ids: set[str] = set()
    unique_teams = []
    for t in teams:
        key = t.external_id or str(id(t))
        if key not in seen_ext_ids:
            seen_ext_ids.add(key)
            unique_teams.append(t)
        else:
            skipped += 1
    teams = unique_teams

    # Restore preserved fields onto matching teams before inserting.
    if replace and preserved:
        for t in teams:
            snap = preserved.get(t.external_id or "")
            if snap:
                if snap["ai_scores"]:
                    t.ai_scores = snap["ai_scores"]
                if snap["judge_scores"]:
                    t.judge_scores = snap["judge_scores"]
                t.advanced_to_round = snap["advanced_to_round"]
                t.final_position = snap["final_position"]
                t.presentation_uploaded = snap["presentation_uploaded"]
                t.repo_url = snap["repo_url"]
                t.repo_ready = snap["repo_ready"]
                t.repo_check_notes = snap["repo_check_notes"]
                t.has_teams_channel = snap["has_teams_channel"]
                t.teams_channel_id = snap["teams_channel_id"]
                t.teams_channel_created_at = snap["teams_channel_created_at"]

    for t in teams:
        db.add(t)
    db.flush()

    summary = screen_all(db)
    db.commit()

    return UploadResult(
        teams_imported=len(teams),
        teams_skipped=skipped,
        duplicate_participants=summary["duplicate_participants"],
        multi_team_mentors=summary["multi_team_mentors"],
    )


@app.post("/api/rescreen", response_model=dict)
def rescreen(db: Session = Depends(get_db)) -> dict:
    summary = screen_all(db)
    db.commit()
    return summary


# ---- AI screening — async background job ----
#
# Bulk AI screening hits the LLM 48× back-to-back and easily blows past
# corporate-LB / nginx timeouts on a single HTTP request. We run it in a
# background thread instead: POST kicks it off and returns a job_id; the
# UI polls GET /api/ai-screen/status every couple of seconds until done.
#
# Single-worker assumption: the systemd unit runs uvicorn with --workers 2,
# so this in-memory state is per-process. Concurrent screening jobs in
# practice never happen (one organizer clicks the button), so the small
# race window if a status check lands on the other worker is acceptable —
# worst case the UI sees stale "idle" for one poll, then catches up.
_ai_screen_logger = logging.getLogger("ai_screen")
# RLock so the same thread can re-enter (e.g. if the handler builds an error
# payload via _ai_screen_status_payload() while still holding the lock).
_ai_screen_lock = threading.RLock()
_ai_screen_state: dict[str, Any] = {
    "running": False,
    "job_id": None,
    "force": False,
    "total": 0,
    "scored": 0,
    "failed": 0,
    "providers": {},
    "started_at": None,
    "finished_at": None,
    "error": None,
}


def _ai_screen_status_payload() -> dict:
    with _ai_screen_lock:
        state = dict(_ai_screen_state)
    if state.get("error"):
        status = "error"
    elif state.get("running"):
        status = "running"
    elif state.get("finished_at"):
        status = "done"
    else:
        status = "idle"
    return {
        "job_id": state.get("job_id"),
        "status": status,
        "total": state.get("total", 0),
        "scored": state.get("scored", 0),
        "failed": state.get("failed", 0),
        "providers": state.get("providers", {}),
        "started_at": state.get("started_at"),
        "finished_at": state.get("finished_at"),
        "error": state.get("error"),
    }


def _run_ai_screen_background(force: bool, job_id: str) -> None:
    """Worker that runs in a background thread.

    Scores teams one at a time so we can publish progress mid-run. Each
    team's score is committed individually so a crash mid-job leaves
    everything before it durably saved.
    """
    try:
        with SessionLocal() as db:
            teams = db.query(models.Team).all()
            scored = 0
            failed = 0
            providers: dict[str, int] = {}
            for t in teams:
                if not force and t.ai_scores and t.ai_scores.get("overall", {}).get("score"):
                    continue
                result = ai_score_team(t)
                t.ai_scores = result
                db.commit()
                if result.get("error"):
                    failed += 1
                else:
                    scored += 1
                    p = result.get("provider") or "unknown"
                    providers[p] = providers.get(p, 0) + 1
                with _ai_screen_lock:
                    _ai_screen_state["scored"] = scored
                    _ai_screen_state["failed"] = failed
                    _ai_screen_state["providers"] = providers
    except Exception as e:  # pragma: no cover — failsafe
        _ai_screen_logger.exception("AI screen background job %s failed", job_id)
        with _ai_screen_lock:
            _ai_screen_state["error"] = str(e)[:200]
    finally:
        with _ai_screen_lock:
            _ai_screen_state["running"] = False
            _ai_screen_state["finished_at"] = datetime.utcnow().isoformat()


@app.post("/api/ai-screen", response_model=dict)
def ai_screen(force: bool = False) -> dict:
    """Start AI screening in the background. Returns immediately with job info.

    Poll GET /api/ai-screen/status to watch progress.
    """
    # Cheap, lockless pre-check — we re-verify under the lock below before mutating
    if _ai_screen_state.get("running"):
        # Build the status payload BEFORE we touch the lock, so we never hold
        # and re-acquire it on the error path.
        existing_status = _ai_screen_status_payload()
        raise HTTPException(
            status_code=409,
            detail={"message": "A screening job is already running", "status": existing_status},
        )

    # Snapshot total team count up-front so the UI has a denominator.
    # This runs outside the lock so a slow DB doesn't block anyone else.
    with SessionLocal() as s:
        total = s.query(models.Team).count()
    job_id = uuid.uuid4().hex[:12]
    started_at = datetime.utcnow().isoformat()

    # Atomic check-then-set: race-safe even if two requests slip past the
    # lockless pre-check at the same time.
    with _ai_screen_lock:
        if _ai_screen_state.get("running"):
            existing_status = _ai_screen_status_payload()
            raise HTTPException(
                status_code=409,
                detail={"message": "A screening job is already running", "status": existing_status},
            )
        _ai_screen_state.update({
            "running": True,
            "job_id": job_id,
            "force": force,
            "total": total,
            "scored": 0,
            "failed": 0,
            "providers": {},
            "started_at": started_at,
            "finished_at": None,
            "error": None,
        })

    threading.Thread(
        target=_run_ai_screen_background,
        args=(force, job_id),
        daemon=True,
        name=f"ai-screen-{job_id}",
    ).start()

    return _ai_screen_status_payload()


@app.get("/api/ai-screen/status", response_model=dict)
def ai_screen_status() -> dict:
    """Current state of the most recent / in-flight AI screening job."""
    return _ai_screen_status_payload()


# ---- Tournament progression (advance teams between rounds, crown winners) ----

@app.post("/api/rounds/advance", response_model=dict)
def advance_round(req: RoundAdvanceRequest, db: Session = Depends(get_db)) -> dict:
    """Mark the listed teams as advanced past from_round into round from_round+1.

    Teams NOT in the list keep their current advanced_to_round — they're
    eliminated by being left behind, not by an explicit demotion. This means
    you can call this endpoint multiple times safely (idempotent for the
    same team_ids; cumulative across different calls).
    """
    if req.from_round not in (1, 2, 3):
        raise HTTPException(status_code=400, detail="from_round must be 1, 2, or 3")
    next_round = req.from_round + 1
    teams = db.query(models.Team).filter(models.Team.id.in_(req.team_ids)).all()
    advanced_ids = []
    for t in teams:
        # Bump only if not already at or past the next round (don't accidentally regress)
        if (t.advanced_to_round or 1) < next_round:
            t.advanced_to_round = next_round
            advanced_ids.append(t.id)
    db.commit()
    return {
        "from_round": req.from_round,
        "to_round": next_round,
        "advanced_team_ids": advanced_ids,
        "advanced_count": len(advanced_ids),
    }


@app.post("/api/rounds/reset/{round}", response_model=dict)
def reset_round_advancements(round: int, db: Session = Depends(get_db)) -> dict:
    """Undo: drop every team's advanced_to_round back to `round` if currently higher.

    Useful when the organizer wants to re-pick the advancing teams for a round.
    Doesn't touch final_position — that's reset via a separate winners call.
    """
    if round not in (1, 2, 3):
        raise HTTPException(status_code=400, detail="round must be 1, 2, or 3")
    teams = db.query(models.Team).filter(models.Team.advanced_to_round > round).all()
    for t in teams:
        t.advanced_to_round = round
    db.commit()
    return {"reset_to_round": round, "teams_reset": len(teams)}


@app.post("/api/rounds/winners", response_model=dict)
def set_winners(req: WinnersSetRequest, db: Session = Depends(get_db)) -> dict:
    """Set 1st / 2nd / 3rd place team IDs. Pass empty positions {} to clear all."""
    # Validate positions
    for pos in req.positions.keys():
        if pos not in ("1", "2", "3"):
            raise HTTPException(status_code=400, detail=f"Position must be '1', '2', or '3'; got {pos}")

    # Clear existing winners first
    for t in db.query(models.Team).filter(models.Team.final_position.isnot(None)).all():
        t.final_position = None

    # Set new winners
    for pos_str, team_id in req.positions.items():
        team = db.query(models.Team).filter(models.Team.id == team_id).first()
        if not team:
            raise HTTPException(status_code=404, detail=f"Team {team_id} not found")
        team.final_position = int(pos_str)
        # Make sure they're at the winner level too
        team.advanced_to_round = max(team.advanced_to_round or 1, 4)
    db.commit()
    return {"positions": req.positions}


@app.get("/api/rounds/summary", response_model=list[RoundSummary])
def round_summary(db: Session = Depends(get_db)) -> list[RoundSummary]:
    """Per-round counts: how many teams eligible, how many already scored."""
    out: list[RoundSummary] = []
    for r in (1, 2, 3):
        eligible = db.query(models.Team).filter(models.Team.advanced_to_round >= r).count()
        scored = (
            db.query(models.JudgeScore.team_id)
            .filter(models.JudgeScore.round == r)
            .distinct()
            .count()
        )
        out.append(RoundSummary(round=r, eligible_team_count=eligible, scored_team_count=scored))
    return out


@app.post("/api/chat", response_model=ChatResponse)
def chat_endpoint(req: ChatRequest, db: Session = Depends(get_db)) -> ChatResponse:
    """Organizer Q&A chatbot over the team dataset.

    Accepts a list of messages (full conversation context) and returns the
    assistant's next reply plus any team IDs the answer references.
    """
    messages = [{"role": m.role, "content": m.content} for m in req.messages]
    result = chat_service.chat(db, messages)
    return ChatResponse(**result)


@app.post("/api/ai-screen/{team_id}", response_model=dict)
def ai_screen_one(team_id: int, db: Session = Depends(get_db)) -> dict:
    team = db.query(models.Team).get(team_id)
    if not team:
        raise HTTPException(status_code=404, detail="team not found")
    result = ai_score_team(team)
    team.ai_scores = result
    db.commit()
    return result


@app.get("/api/email/templates", response_model=list[EmailTemplateOut])
def list_email_templates() -> list[EmailTemplateOut]:
    # Archived templates (mentor_confirm, final_call, individual_participation
    # for 2026 — registration closed) are kept in source so they can be
    # un-archived for the next event, but hidden from the composer dropdown.
    return [
        EmailTemplateOut(id=t.id, label=t.label, description=t.description, audience=t.audience, subject=t.subject)
        for t in EMAIL_TEMPLATES
        if not t.archived
    ]


@app.post("/api/email/render", response_model=list[RenderedEmail])
def render_email_endpoint(req: EmailRenderRequest, db: Session = Depends(get_db)) -> list[RenderedEmail]:
    try:
        rendered = render_emails(db, req.template_id, req.team_ids)
    except KeyError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return [RenderedEmail(**r) for r in rendered]


@app.get("/api/email/templates/{template_id}/blank", response_model=dict)
def render_template_blank(template_id: str) -> dict:
    """Return a template's subject + body + body_html with any
    {placeholder} tokens stripped out, so it can be sent to arbitrary
    recipients (judges, observers, SLT) without picking a team first.

    Designed for the Kickoff-style email which uses a generic 'Dear Team'
    greeting and has no per-team substitution. For templates that DO
    have placeholders (welcome, fix_it, channel_ready), this returns
    the body with the placeholders removed -- still ok if the surrounding
    text reads naturally without them; otherwise the organizer should
    use the team-based flow.
    """
    import re
    template = next((t for t in EMAIL_TEMPLATES if t.id == template_id), None)
    if template is None:
        raise HTTPException(status_code=404, detail="unknown template id")

    placeholder = re.compile(r"\{[a-zA-Z_][a-zA-Z0-9_]*\}")
    return {
        "template_id": template_id,
        "subject": placeholder.sub("", template.subject).strip(),
        "body": placeholder.sub("", template.body),
        "body_html": placeholder.sub("", template.body_html) if template.body_html else None,
    }


@app.post("/api/judge/{team_id}/ai", response_model=dict)
def judge_ai_endpoint(team_id: int, req: JudgeAIRequest, db: Session = Depends(get_db)) -> dict:
    team = db.query(models.Team).get(team_id)
    if not team:
        raise HTTPException(status_code=404, detail="team not found")
    if req.repo_url is not None:
        team.repo_url = req.repo_url or None
    result = ai_judge(team, github_url=req.repo_url or team.repo_url)
    existing = team.judge_scores or {}
    existing["ai"] = result["ai"]
    existing["github"] = result["github"]
    team.judge_scores = existing
    db.commit()
    return existing


@app.get("/api/export.csv")
def export_csv(db: Session = Depends(get_db)) -> StreamingResponse:
    teams = db.query(models.Team).order_by(models.Team.name.asc()).all()
    buf = io.StringIO()
    w = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    w.writerow([
        "team_name", "mentor_name", "mentor_email",
        "members", "locations", "tshirt_sizes",
        "completeness_pct", "flag_count", "flags",
        "ai_overall", "ai_genuineness", "ai_solution_clarity", "ai_business_value", "ai_novelty",
        "judge_ai_overall", "judge_human_overall", "judge_panelist",
        "idea", "tools", "approach", "viability", "business_value",
        "repo_url",
    ])
    for t in teams:
        ai = t.ai_scores or {}
        js = t.judge_scores or {}
        j_ai = js.get("ai", {}) or {}
        j_h = js.get("human", {}) or {}
        w.writerow([
            t.name,
            t.mentor_name or "",
            t.mentor_email or "",
            " | ".join(m.name for m in t.members),
            " | ".join((m.location or "") for m in t.members),
            " | ".join((m.tshirt_size or "") for m in t.members),
            int(round((t.completeness_score or 0) * 100)),
            len(t.flags or []),
            " | ".join(t.flags or []),
            (ai.get("overall") or {}).get("score", ""),
            (ai.get("genuineness") or {}).get("score", ""),
            (ai.get("solution_clarity") or {}).get("score", ""),
            (ai.get("business_value") or {}).get("score", ""),
            (ai.get("novelty") or {}).get("score", ""),
            j_ai.get("overall", "") if not j_ai.get("error") else "",
            j_h.get("overall", "") or "",
            j_h.get("panelist") or "",
            (t.idea or "").replace("\n", " ").strip(),
            (t.tools or "").replace("\n", " ").strip(),
            (t.approach or "").replace("\n", " ").strip(),
            (t.viability or "").replace("\n", " ").strip(),
            (t.business_value or "").replace("\n", " ").strip(),
            t.repo_url or "",
        ])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="realhack_pilot_export.csv"'},
    )


@app.get("/api/export/devops-repos.xlsx")
def export_devops_repos(db: Session = Depends(get_db)) -> StreamingResponse:
    """Hand-off sheet for DevOps to provision GitHub repos.

    One row per person (mentor first, then each member). S.No, Team Name and
    Gitrepo url cells are merged vertically across the team's rows so the
    DevOps team sees one team-grouped block at a glance, matching the layout
    organizers asked for. Gitrepo url is left blank for DevOps to fill in
    (or pre-filled if we already have one in repo_url).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    teams = db.query(models.Team).order_by(models.Team.name.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "DevOps Repos"

    headers = ["S.No", "Team Name", "Member names", "Email Ids", "Gitrepo url"]
    header_fill = PatternFill("solid", fgColor="0A4F99")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="B0BEC5")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border

    row = 2
    for sno, t in enumerate(teams, start=1):
        # Mentor row first, then each member. Skip teams with neither.
        people: list[tuple[str, str]] = []
        if (t.mentor_name or "").strip() or (t.mentor_email or "").strip():
            people.append((
                (t.mentor_name or "").strip() or "(mentor name missing)",
                (t.mentor_email or "").strip(),
            ))
        for m in t.members:
            people.append((
                (m.name or "").strip(),
                (m.email or "").strip(),
            ))
        if not people:
            continue

        start_row = row
        for name, email in people:
            ws.cell(row=row, column=3, value=name).alignment = left
            ws.cell(row=row, column=4, value=email).alignment = left
            for col_idx in range(1, 6):
                ws.cell(row=row, column=col_idx).border = border
            row += 1
        end_row = row - 1

        # Write the team-level cells in the first row of the block; merge.
        ws.cell(row=start_row, column=1, value=sno).alignment = center
        ws.cell(row=start_row, column=2, value=t.name).alignment = center
        ws.cell(row=start_row, column=5, value=(t.repo_url or "")).alignment = center
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, end_row=end_row, start_column=1, end_column=1)
            ws.merge_cells(start_row=start_row, end_row=end_row, start_column=2, end_column=2)
            ws.merge_cells(start_row=start_row, end_row=end_row, start_column=5, end_column=5)

    # Column widths sized for the longest realistic value in each column.
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 42
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="realhack_2026_devops_repos.xlsx"'},
    )


@app.post("/api/judge/{team_id}/human", response_model=dict)
def judge_human_endpoint(team_id: int, req: JudgeHumanRequest, db: Session = Depends(get_db)) -> dict:
    team = db.query(models.Team).get(team_id)
    if not team:
        raise HTTPException(status_code=404, detail="team not found")
    if req.repo_url is not None:
        team.repo_url = req.repo_url or None
    existing = merge_human_scores(team.judge_scores or {}, req.scores, req.panelist)
    team.judge_scores = existing
    db.commit()
    return existing


# ===== Live judge panel scoring (new M4) =====

@app.get("/api/judging/rubric")
def judging_rubric() -> dict:
    return {"axes": [{"key": k, "label": v} for k, v in JUDGE_RUBRIC_AXES], "max_per_axis": 10}


@app.get("/api/judges", response_model=list[JudgeOut])
def list_judges(db: Session = Depends(get_db)) -> list[models.Judge]:
    return db.query(models.Judge).filter(models.Judge.is_active == True).order_by(models.Judge.name.asc()).all()  # noqa: E712


@app.post("/api/judges", response_model=JudgeOut)
def create_judge(req: JudgeCreate, db: Session = Depends(get_db)) -> models.Judge:
    judge = judging.upsert_judge_by_email(db, name=req.name, email=req.email, role=req.role)
    db.commit()
    return judge


@app.post("/api/judges/bulk", response_model=JudgeBulkResult)
def bulk_add_judges(req: JudgeBulkRequest, db: Session = Depends(get_db)) -> JudgeBulkResult:
    """Bulk upsert judges by email. Skips no-op duplicates (same name + role
    already present) so re-running is safe. Returns per-row counts + any
    rows that failed validation."""
    created = 0
    updated = 0
    skipped = 0
    failed: list[dict] = []

    for row in req.rows:
        name = (row.name or "").strip()
        email = (row.email or "").strip().lower()
        role = (row.role or "judge").strip().lower()
        if role not in ("judge", "organizer", "rews"):
            failed.append({"name": name, "email": email, "error": "role must be 'judge', 'organizer', or 'rews'"})
            continue
        if not name or not email or "@" not in email:
            failed.append({"name": name, "email": email, "error": "name and email required (and email must look like one)"})
            continue
        # Case-insensitive match against existing rows — older rows stored
        # via MSAL login may have mixed-case emails; we must find them or
        # we'll create a duplicate row for the same person.
        existing = db.query(models.Judge).filter(func.lower(models.Judge.email) == email).first()
        if existing:
            changed = False
            if existing.name != name:
                existing.name = name
                changed = True
            if existing.role != role:
                existing.role = role
                changed = True
            if not existing.is_active:
                existing.is_active = True
                changed = True
            if changed:
                updated += 1
            else:
                skipped += 1
        else:
            judging.upsert_judge_by_email(db, name=name, email=email, role=role)
            created += 1

    db.commit()
    return JudgeBulkResult(
        created_count=created,
        updated_count=updated,
        skipped_count=skipped,
        failed=failed,
    )


@app.post("/api/judges/dedupe-emails", response_model=dict)
def dedupe_judge_emails(db: Session = Depends(get_db)) -> dict:  # noqa: C901 — single linear merge loop, deliberately not split
    try:
        return _dedupe_judge_emails_impl(db)
    except HTTPException:
        raise
    except Exception as e:
        # Bubble the actual error message back to the UI so the user (and
        # next-me reading the toast) doesn't have to SSH to the server to
        # find out what broke. 500 keeps the response code accurate.
        import traceback
        tb = traceback.format_exc(limit=4)
        raise HTTPException(status_code=500, detail=f"Dedupe failed: {e!s} | {tb[-400:]}")


def _dedupe_judge_emails_impl(db: Session) -> dict:
    """Merge judge rows that share the same email modulo case, then lowercase
    every remaining email. Idempotent — running it after the data is clean
    returns merged_count: 0.

    For each duplicate group:
      - Picks a 'canonical' row to keep: prefers organizer > judge, then rows
        with attached judge_scores > rows with panel memberships > oldest id.
        This preserves any work the user has already done (scoring, panel
        assignment) against the canonical row.
      - Reassigns FK references (panel_judges.judge_id, judge_score_records.
        judge_id, judge_assignments.judge_id) from each duplicate to canonical,
        respecting the unique-constraint on (judge_id, ...) by dropping the
        duplicate's row if canonical already has the same composite key.
      - Deletes the duplicate Judge rows.
      - Lowercases the canonical row's email so future equality checks line up.
    """
    # Pull all judges grouped by lowercase email.
    rows = db.query(models.Judge).all()
    by_email: dict[str, list[models.Judge]] = {}
    for j in rows:
        key = (j.email or "").strip().lower()
        if not key:
            continue
        by_email.setdefault(key, []).append(j)

    merged_count = 0
    deleted_ids: list[int] = []

    def score_judge(j: models.Judge) -> tuple:
        # Higher tuple wins. Organizer > judge; rows with data > empty rows.
        score_count = db.query(models.JudgeScore).filter(models.JudgeScore.judge_id == j.id).count()
        panel_count = db.query(models.PanelJudge).filter(models.PanelJudge.judge_id == j.id).count()
        is_org = 1 if (j.role == "organizer") else 0
        # Prefer oldest id last (smaller id wins → invert with -id).
        return (is_org, score_count, panel_count, -j.id)

    for email, dups in by_email.items():
        if len(dups) < 2:
            # Even singletons get their email lowercased for consistency.
            if dups[0].email != email:
                dups[0].email = email
            continue

        dups.sort(key=score_judge, reverse=True)
        canonical = dups[0]
        losers = dups[1:]

        # CRITICAL: clear canonical's email FIRST so the unique constraint on
        # `judges.email` doesn't fire when SQLAlchemy reorders the flush. Both
        # the loser DELETE and the canonical UPDATE (to lowercased email) are
        # queued — without this, the UPDATE can hit the constraint before the
        # DELETE of the row holding the lowercased variant lands, causing a
        # 500. We restore the lowercased email at the end after all deletes
        # are flushed.
        canonical_target_email = email
        canonical.email = f"__dedupe_pending__{canonical.id}@invalid.local"
        db.flush()

        for loser in losers:
            # judge_score_records: unique on (judge_id, team_id, round)
            for s in db.query(models.JudgeScore).filter(models.JudgeScore.judge_id == loser.id).all():
                clash = db.query(models.JudgeScore).filter(
                    models.JudgeScore.judge_id == canonical.id,
                    models.JudgeScore.team_id == s.team_id,
                    models.JudgeScore.round == s.round,
                ).first()
                if clash:
                    db.delete(s)
                else:
                    s.judge_id = canonical.id

            # panel_judges: unique on (panel_id, judge_id)
            for pj in db.query(models.PanelJudge).filter(models.PanelJudge.judge_id == loser.id).all():
                clash = db.query(models.PanelJudge).filter(
                    models.PanelJudge.panel_id == pj.panel_id,
                    models.PanelJudge.judge_id == canonical.id,
                ).first()
                if clash:
                    db.delete(pj)
                else:
                    pj.judge_id = canonical.id

            # judge_assignments (legacy): unique on (judge_id, team_id, round)
            for ja in db.query(models.JudgeAssignment).filter(models.JudgeAssignment.judge_id == loser.id).all():
                clash = db.query(models.JudgeAssignment).filter(
                    models.JudgeAssignment.judge_id == canonical.id,
                    models.JudgeAssignment.team_id == ja.team_id,
                    models.JudgeAssignment.round == ja.round,
                ).first()
                if clash:
                    db.delete(ja)
                else:
                    ja.judge_id = canonical.id

            db.flush()
            deleted_ids.append(loser.id)
            db.delete(loser)
            merged_count += 1

        # Force all loser DELETEs to land before we reclaim the target email.
        db.flush()
        canonical.email = canonical_target_email
        db.flush()

    db.commit()
    return {
        "merged_count": merged_count,
        "deleted_ids": deleted_ids,
        "remaining_judges": db.query(models.Judge).count(),
    }


# Built-in organizers that can never be removed — guards the core team from
# accidental deletion. Stored lowercase for case-insensitive compares.
PROTECTED_EMAILS: set[str] = {
    "shaikshavali.kalluri@realpage.com",
    "suneel.nallu@realpage.com",
    "bhaskar.jaddu@realpage.com",
}


@app.patch("/api/judges/{judge_id}", response_model=JudgeOut)
def update_judge(judge_id: int, req: JudgeUpdate, db: Session = Depends(get_db)) -> models.Judge:
    judge = db.query(models.Judge).get(judge_id)
    if not judge:
        raise HTTPException(status_code=404, detail="judge not found")
    if req.name and req.name.strip():
        judge.name = req.name.strip()
    if req.email is not None:
        new_email = (req.email or "").strip().lower() or None
        # Protected accounts: don't let their email be changed (which would lock them out).
        if (judge.email or "").strip().lower() in PROTECTED_EMAILS and new_email != (judge.email or "").strip().lower():
            raise HTTPException(status_code=403, detail="Protected account email cannot be changed")
        judge.email = new_email
    if req.role:
        new_role = req.role.strip().lower()
        if new_role not in ("judge", "organizer", "rews"):
            raise HTTPException(status_code=400, detail="role must be 'judge', 'organizer', or 'rews'")
        # Protected accounts cannot be downgraded from organizer.
        if (judge.email or "").strip().lower() in PROTECTED_EMAILS and new_role != "organizer":
            raise HTTPException(status_code=403, detail="Protected account must remain an organizer")
        judge.role = new_role
    db.commit()
    db.refresh(judge)
    return judge


@app.delete("/api/judges/{judge_id}")
def delete_judge(judge_id: int, db: Session = Depends(get_db)) -> dict:
    judge = db.query(models.Judge).get(judge_id)
    if not judge:
        raise HTTPException(status_code=404, detail="judge not found")
    if (judge.email or "").strip().lower() in PROTECTED_EMAILS:
        raise HTTPException(status_code=403, detail="This account is protected and cannot be deleted")
    db.delete(judge)
    db.commit()
    return {"deleted": True}


@app.post("/api/judges/login", response_model=JudgeOut)
def judge_login(req: JudgeCreate, db: Session = Depends(get_db)) -> models.Judge:
    """Mock SSO: upsert by email, return Judge record."""
    judge = judging.upsert_judge_by_email(db, name=req.name, email=req.email, role=req.role or "judge")
    db.commit()
    return judge


@app.get("/api/me/role", response_model=UserRoleOut)
def me_role(
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
    db: Session = Depends(get_db),
) -> UserRoleOut:
    """After Azure AD login succeeds, look up the user's role in our Judge table.

    Azure AD + AD group already guard the API at the network/auth layer — this
    endpoint just maps the signed-in user's email to organizer/judge/none so
    the frontend can route them to the right view.
    """
    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    email = (profile.get("email") or "").strip().lower()
    if not email:
        return UserRoleOut(role="none", name=profile.get("name"), email=None)
    judge = (
        db.query(models.Judge)
        .filter(models.Judge.email.ilike(email))
        .filter(models.Judge.is_active == True)  # noqa: E712
        .first()
    )
    if not judge:
        # Super-admin always passes — needed so Shaik can still access the app
        # in Test Mode before the sandbox DB has been populated from prod.
        if email in SANDBOX_ADMIN_EMAILS:
            return UserRoleOut(
                role="organizer",
                judge_id=None,
                name=profile.get("name"),
                email=email,
            )
        # Anyone signed in but not on the roster: tell the frontend to show
        # the 'Not registered' card. REWS volunteers should be added to the
        # judges table with role='rews' explicitly so they get the scoped
        # pickup-desk UI -- we don't want implicit access by virtue of
        # 'signed in but not on the roster'.
        return UserRoleOut(role="none", name=profile.get("name"), email=email)
    return UserRoleOut(
        role=judge.role or "judge",
        judge_id=judge.id,
        name=judge.name,
        email=judge.email,
    )


def _available_rounds_for_judge(db: Session, judge_id: int) -> list[int]:
    """Rounds where this judge has at least one panel that already has both
    teams AND judges (i.e. the round is actually ready to be scored)."""
    panel_rows = db.query(models.Panel).join(
        models.PanelJudge, models.PanelJudge.panel_id == models.Panel.id
    ).filter(models.PanelJudge.judge_id == judge_id).all()
    rounds: set[int] = set()
    for p in panel_rows:
        # require at least one team in the panel — otherwise round is effectively empty
        if any(True for _ in p.teams):
            rounds.add(p.round)
    return sorted(rounds)


def _team_ids_for_judge_via_panels(db: Session, judge_id: int, round: int | None) -> set[int]:
    """Return the set of team ids the judge sees, derived from panel membership.

    A judge sees every team that's in at least one panel they're a member of
    (filtered by round if supplied). Same team across multiple panels is
    deduped naturally by the set.
    """
    pj_q = db.query(models.PanelJudge).filter(models.PanelJudge.judge_id == judge_id)
    panel_ids = [pj.panel_id for pj in pj_q.all()]
    if not panel_ids:
        return set()
    panels_q = db.query(models.Panel).filter(models.Panel.id.in_(panel_ids))
    if round is not None:
        panels_q = panels_q.filter(models.Panel.round == round)
    relevant_panel_ids = [p.id for p in panels_q.all()]
    if not relevant_panel_ids:
        return set()
    rows = db.query(models.PanelTeam).filter(models.PanelTeam.panel_id.in_(relevant_panel_ids)).all()
    return {r.team_id for r in rows}


@app.get("/api/judge/me/teams", response_model=list[TeamOut])
def my_assigned_teams(
    round: int | None = None,
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
    db: Session = Depends(get_db),
) -> list[models.Team]:
    """Teams the signed-in judge sees in the given round, via panel membership."""
    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    email = (profile.get("email") or "").strip().lower()
    judge = (
        db.query(models.Judge)
        .filter(models.Judge.email.ilike(email))
        .filter(models.Judge.is_active == True)  # noqa: E712
        .first()
    )
    if not judge:
        raise HTTPException(status_code=403, detail="Not registered as a judge")
    team_ids = _team_ids_for_judge_via_panels(db, judge.id, round)
    if not team_ids:
        return []
    return db.query(models.Team).filter(models.Team.id.in_(team_ids)).order_by(models.Team.name.asc()).all()


@app.get("/api/judge/me/rounds", response_model=list[int])
def my_available_rounds(
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
    db: Session = Depends(get_db),
) -> list[int]:
    """Rounds the signed-in judge can score in (i.e. has panels with teams)."""
    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    email = (profile.get("email") or "").strip().lower()
    judge = (
        db.query(models.Judge)
        .filter(models.Judge.email.ilike(email))
        .filter(models.Judge.is_active == True)  # noqa: E712
        .first()
    )
    if not judge:
        raise HTTPException(status_code=403, detail="Not registered as a judge")
    return _available_rounds_for_judge(db, judge.id)


@app.get("/api/judges/{judge_id}/rounds", response_model=list[int])
def rounds_for_judge(judge_id: int, db: Session = Depends(get_db)) -> list[int]:
    """Rounds a specific judge can score in. Used by 'preview as judge'."""
    return _available_rounds_for_judge(db, judge_id)


@app.get("/api/judges/{judge_id}/teams", response_model=list[TeamOut])
def teams_for_judge(
    judge_id: int,
    round: int | None = None,
    db: Session = Depends(get_db),
) -> list[models.Team]:
    """Teams the given judge sees in the given round (via panel membership).
    Used by the organizer 'preview as judge' feature."""
    team_ids = _team_ids_for_judge_via_panels(db, judge_id, round)
    if not team_ids:
        return []
    return db.query(models.Team).filter(models.Team.id.in_(team_ids)).order_by(models.Team.name.asc()).all()


# ===== Panels: groups of teams + judges per round =====

def _panel_to_out(p: models.Panel) -> PanelOut:
    return PanelOut(
        id=p.id,
        name=p.name,
        round=p.round,
        team_ids=sorted({pt.team_id for pt in p.teams}),
        judge_ids=sorted({pj.judge_id for pj in p.judges}),
    )


@app.get("/api/panels", response_model=list[PanelOut])
def list_panels(round: int | None = None, db: Session = Depends(get_db)) -> list[PanelOut]:
    q = db.query(models.Panel)
    if round is not None:
        q = q.filter(models.Panel.round == round)
    panels = q.order_by(models.Panel.round.asc(), models.Panel.id.asc()).all()
    return [_panel_to_out(p) for p in panels]


@app.post("/api/panels", response_model=PanelOut)
def create_panel(
    req: PanelCreate,
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
    db: Session = Depends(get_db),
) -> PanelOut:
    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    panel = models.Panel(
        name=req.name.strip() or f"Panel {req.round}",
        round=req.round,
        created_by_email=(profile.get("email") or "").lower() or None,
    )
    db.add(panel)
    db.commit()
    db.refresh(panel)
    return _panel_to_out(panel)


@app.patch("/api/panels/{panel_id}", response_model=PanelOut)
def update_panel(panel_id: int, req: PanelUpdate, db: Session = Depends(get_db)) -> PanelOut:
    panel = db.query(models.Panel).get(panel_id)
    if not panel:
        raise HTTPException(status_code=404, detail="panel not found")
    if req.name is not None and req.name.strip():
        panel.name = req.name.strip()
    db.commit()
    db.refresh(panel)
    return _panel_to_out(panel)


@app.delete("/api/panels/{panel_id}")
def delete_panel(panel_id: int, db: Session = Depends(get_db)) -> dict:
    panel = db.query(models.Panel).get(panel_id)
    if not panel:
        raise HTTPException(status_code=404, detail="panel not found")
    db.delete(panel)
    db.commit()
    return {"deleted": True}


@app.post("/api/panels/{panel_id}/teams", response_model=PanelOut)
def set_panel_teams(panel_id: int, req: PanelTeamsSet, db: Session = Depends(get_db)) -> PanelOut:
    """Replace the set of teams in this panel."""
    panel = db.query(models.Panel).get(panel_id)
    if not panel:
        raise HTTPException(status_code=404, detail="panel not found")
    db.query(models.PanelTeam).filter(models.PanelTeam.panel_id == panel_id).delete(synchronize_session=False)
    for tid in set(req.team_ids):
        if not db.query(models.Team).get(tid):
            continue
        db.add(models.PanelTeam(panel_id=panel_id, team_id=tid))
    db.commit()
    db.refresh(panel)
    return _panel_to_out(panel)


@app.post("/api/panels/{panel_id}/judges", response_model=PanelOut)
def set_panel_judges(panel_id: int, req: PanelJudgesSet, db: Session = Depends(get_db)) -> PanelOut:
    """Replace the set of judges in this panel."""
    panel = db.query(models.Panel).get(panel_id)
    if not panel:
        raise HTTPException(status_code=404, detail="panel not found")
    db.query(models.PanelJudge).filter(models.PanelJudge.panel_id == panel_id).delete(synchronize_session=False)
    for jid in set(req.judge_ids):
        if not db.query(models.Judge).get(jid):
            continue
        db.add(models.PanelJudge(panel_id=panel_id, judge_id=jid))
    db.commit()
    db.refresh(panel)
    return _panel_to_out(panel)


@app.get("/api/panels/{panel_id}/invite.ics")
def get_panel_invite_ics(
    panel_id: int,
    day: int,
    claims: dict = Depends(require_auth),
    db: Session = Depends(get_db),
) -> Response:
    """Generate an Outlook-ready .ics invite for a panel's judging day.

    Day 1 = June 18, Day 2 = June 19, 2026 (9:00-17:00 IST, 13:00-14:00 lunch
    excluded, 15-min slots). The signed-in user is set as ORGANIZER so when
    they open the file Outlook shows Send rather than Accept/Decline.
    """
    if day not in (1, 2):
        raise HTTPException(status_code=400, detail="day must be 1 or 2")
    panel = db.query(models.Panel).get(panel_id)
    if not panel:
        raise HTTPException(status_code=404, detail="panel not found")

    organizer_email = (claims.get("preferred_username") or claims.get("upn") or "").strip()
    organizer_name = (claims.get("name") or organizer_email).strip()

    try:
        ics_text = scheduling.build_panel_invite_ics(
            panel,
            day=day,
            organizer_email=organizer_email,
            organizer_name=organizer_name,
            db=db,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    safe_panel_name = "".join(c if c.isalnum() or c in "-_" else "-" for c in panel.name.lower())
    filename = f"realhack-{safe_panel_name}-day{day}.ics"
    return Response(
        content=ics_text,
        media_type="text/calendar; method=REQUEST; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/panels/{panel_id}/invite-meta")
def get_panel_invite_meta(
    panel_id: int,
    day: int,
    claims: dict = Depends(require_auth),
    db: Session = Depends(get_db),
) -> dict:
    """Return JSON meeting metadata for the Outlook-Web compose deeplink path.

    New Outlook doesn't reliably treat downloaded .ics files as editable
    drafts. This endpoint feeds the frontend's clipboard + deeplink flow:
    the frontend copies the attendee emails to the clipboard and opens
    outlook.office.com/calendar/deeplink/compose with subject/start/end/
    body/location, where new Outlook intercepts and opens its native
    Create-event dialog with the user as organizer.
    """
    if day not in (1, 2):
        raise HTTPException(status_code=400, detail="day must be 1 or 2")
    panel = db.query(models.Panel).get(panel_id)
    if not panel:
        raise HTTPException(status_code=404, detail="panel not found")
    try:
        return scheduling.build_panel_invite_meta(panel, day=day, db=db)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/panels/{panel_id}/swap-team-days")
def swap_panel_team_days(
    panel_id: int,
    req: PanelSwapTeamDays,
    claims: dict = Depends(require_auth),
    db: Session = Depends(get_db),
) -> dict:
    """Swap two teams' day assignments in a panel. Both teams must belong
    to this panel. Stores the override in panel_team_day_overrides so the
    decision survives subsequent invite-meta requests.
    """
    panel = db.query(models.Panel).get(panel_id)
    if not panel:
        raise HTTPException(status_code=404, detail="panel not found")
    if req.team_a_id == req.team_b_id:
        raise HTTPException(status_code=400, detail="team_a_id and team_b_id must differ")

    panel_team_ids = {pt.team_id for pt in panel.teams}
    if req.team_a_id not in panel_team_ids or req.team_b_id not in panel_team_ids:
        raise HTTPException(status_code=400, detail="both teams must belong to this panel")

    # Compute the current day for each team using the existing distribution
    # (which already honors any prior overrides).
    all_teams = [pt.team for pt in panel.teams if pt.team is not None]
    existing_overrides = scheduling._load_day_overrides(panel, db)
    day1_teams, day2_teams = scheduling._distribute_teams_across_days(
        all_teams, overrides=existing_overrides
    )
    day_for = {t.id: 1 for t in day1_teams}
    day_for.update({t.id: 2 for t in day2_teams})

    a_day = day_for.get(req.team_a_id)
    b_day = day_for.get(req.team_b_id)
    if a_day is None or b_day is None:
        raise HTTPException(status_code=400, detail="could not resolve current day for one of the teams")
    if a_day == b_day:
        raise HTTPException(
            status_code=400,
            detail=f"both teams are already on Day {a_day} — nothing to swap",
        )

    def upsert(team_id: int, day: int) -> None:
        existing = (
            db.query(models.PanelTeamDayOverride)
            .filter(
                models.PanelTeamDayOverride.panel_id == panel_id,
                models.PanelTeamDayOverride.team_id == team_id,
            )
            .first()
        )
        if existing:
            existing.day = day
        else:
            db.add(models.PanelTeamDayOverride(panel_id=panel_id, team_id=team_id, day=day))

    upsert(req.team_a_id, b_day)  # team A takes team B's day
    upsert(req.team_b_id, a_day)
    db.commit()
    return {"swapped": True, "team_a_id": req.team_a_id, "team_a_new_day": b_day, "team_b_id": req.team_b_id, "team_b_new_day": a_day}


# ===== Swag (t-shirt) pickup =====
#
# Used by organizers at the event-day pickup desk. Replaces the shared
# Excel-tracker workflow: multiple organizers can mark people collected
# concurrently from their phones, no merge conflicts, instant search.


def _normalize_country(raw: str | None) -> str | None:
    """Map common location strings to a canonical country name for the swag UI."""
    if not raw:
        return None
    s = raw.strip().lower()
    if not s:
        return None
    mapping = {
        "us": "US",
        "usa": "US",
        "united states": "US",
        "united states of america": "US",
        "india": "India",
        "in": "India",
        "philippines": "Philippines",
        "ph": "Philippines",
        "uk": "UK",
        "united kingdom": "UK",
        "canada": "Canada",
        "ca": "Canada",
        "romania": "Romania",
        "mexico": "Mexico",
    }
    return mapping.get(s, raw.strip())


def _swag_roster(db: Session) -> list[dict]:
    """Build the canonical pickup list — one entry per unique email across
    all members + mentors + non-team extras (judges/organisers/support/etc.),
    with their team(s), role(s), country, category badge, and collection state.
    """
    by_email: dict[str, dict] = {}

    def _add(email: str | None, name: str | None, tshirt_size: str | None, country: str | None, role: str, team_name: str) -> None:
        if not email:
            return
        key = email.strip().lower()
        if not key:
            return
        if key not in by_email:
            by_email[key] = {
                "email": key,
                "name": (name or "").strip() or key,
                "tshirt_size": tshirt_size,
                "country": _normalize_country(country),
                "roles": [],
                "teams": [],
                "category": "Mentor" if role == "mentor" else "Member",
            }
        entry = by_email[key]
        entry["roles"].append(f"{role}:{team_name}")
        if team_name not in entry["teams"]:
            entry["teams"].append(team_name)
        # Mentor role outranks member -- if the same email appears as both
        # (rare but possible), show 'Mentor' on the badge.
        if role == "mentor":
            entry["category"] = "Mentor"
        # Keep first non-empty values seen (member + mentor rows may have
        # different completeness; trust whatever appeared first).
        if not entry["tshirt_size"] and tshirt_size:
            entry["tshirt_size"] = tshirt_size
        if not entry["country"] and country:
            entry["country"] = _normalize_country(country)

    for team in db.query(models.Team).all():
        for m in team.members:
            _add(m.email, m.name, m.tshirt_size, m.location, "member", team.name)
        _add(team.mentor_email, team.mentor_name, team.mentor_tshirt_size, team.mentor_location, "mentor", team.name)

    # Merge in the non-team extras (Judges, Organisers, Support, HR, ...).
    # If an extra's email already appears as a member/mentor, the team
    # role wins on the badge -- they're getting the t-shirt either way,
    # and 'Mentor of TeamX' is more useful at pickup than 'Judge'. The
    # extras entry then just supplements anything that's missing (size,
    # country) without changing the category.
    try:
        extras_rows = db.query(models.SwagExtra).all()
    except Exception:
        # Table may not exist yet (fresh DB, ownership not transferred).
        # Don't crash the roster fetch -- just return the member/mentor list.
        extras_rows = []
    for x in extras_rows:
        key = (x.email or "").strip().lower()
        if not key:
            continue
        if key in by_email:
            entry = by_email[key]
            if not entry.get("tshirt_size") and x.tshirt_size:
                entry["tshirt_size"] = x.tshirt_size
            if not entry.get("country") and x.country:
                entry["country"] = _normalize_country(x.country)
            continue
        by_email[key] = {
            "email": key,
            "name": (x.name or "").strip() or key,
            "tshirt_size": x.tshirt_size,
            "country": _normalize_country(x.country),
            "roles": [f"{(x.category or 'extra').lower()}:"],
            "teams": [],
            "category": (x.category or "Extra").strip() or "Extra",
        }

    # Merge in collection state from swag_pickups
    pickups = {p.email.lower(): p for p in db.query(models.SwagPickup).all()}
    for key, entry in by_email.items():
        p = pickups.get(key)
        entry["collected"] = p is not None
        entry["collected_at"] = p.collected_at.isoformat() if p else None
        entry["collected_by_email"] = p.collected_by_email if p else None
        entry["picked_up_by_name"] = p.picked_up_by_name if p else None
        entry["picked_up_by_email"] = p.picked_up_by_email if p else None
        entry["notes"] = p.notes if p else None
        # Normalize size capitalization
        if entry["tshirt_size"]:
            entry["tshirt_size"] = entry["tshirt_size"].upper()

    return list(by_email.values())


@app.get("/api/swag/people", response_model=list[SwagPersonOut])
def list_swag_people(db: Session = Depends(get_db)) -> list[SwagPersonOut]:
    """Full pickup list — every unique participant + mentor with their
    t-shirt size, team(s), and collection state. Sorted alphabetically.
    """
    rows = _swag_roster(db)
    rows.sort(key=lambda r: r["name"].lower())
    return [SwagPersonOut(**r) for r in rows]


@app.get("/api/swag/stats", response_model=SwagStats)
def swag_stats(db: Session = Depends(get_db)) -> SwagStats:
    rows = _swag_roster(db)
    total = len(rows)
    collected = sum(1 for r in rows if r["collected"])
    return SwagStats(total=total, collected=collected, pending=total - collected)


@app.post("/api/swag/mark", response_model=SwagPersonOut)
def mark_swag_collected(
    req: SwagMarkRequest,
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
    db: Session = Depends(get_db),
) -> SwagPersonOut:
    """Mark a person's t-shirt as collected. Audit-logged with the organizer
    who clicked the button. Idempotent — clicking again on an already-collected
    row just updates the notes and timestamp without erroring.
    """
    email = req.email.strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="email is required")
    # Confirm this email is actually on the pickup roster (member or mentor)
    roster = {r["email"]: r for r in _swag_roster(db)}
    if email not in roster:
        raise HTTPException(status_code=404, detail=f"No participant/mentor on file with email {email}")

    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    organizer_email = (profile.get("email") or "").strip().lower() or None

    pickup = db.query(models.SwagPickup).filter(models.SwagPickup.email == email).first()
    # Normalize on-behalf-of fields (treat empty strings as None)
    pby_name = (req.picked_up_by_name or "").strip() or None
    pby_email = (req.picked_up_by_email or "").strip().lower() or None
    if pickup:
        # Already collected — refresh timestamp + notes, swap in current organizer
        pickup.collected_at = datetime.utcnow()
        pickup.collected_by_email = organizer_email
        pickup.picked_up_by_name = pby_name
        pickup.picked_up_by_email = pby_email
        if req.notes is not None:
            pickup.notes = req.notes
    else:
        person = roster[email]
        pickup = models.SwagPickup(
            email=email,
            person_name=person["name"],
            tshirt_size=person["tshirt_size"],
            collected_at=datetime.utcnow(),
            collected_by_email=organizer_email,
            picked_up_by_name=pby_name,
            picked_up_by_email=pby_email,
            notes=req.notes,
        )
        db.add(pickup)
    db.commit()
    # Refresh roster entry for response
    fresh = next(r for r in _swag_roster(db) if r["email"] == email)
    return SwagPersonOut(**fresh)


@app.post("/api/swag/unmark", response_model=SwagPersonOut)
def unmark_swag_collected(
    req: SwagMarkRequest,
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
    db: Session = Depends(get_db),
) -> SwagPersonOut:
    """Undo a mark — used if an organizer accidentally clicked the wrong person.

    Restricted to **organizers only**. REWS volunteers + judges at the
    desk can mark but not unmark, so a fat-finger at the pickup desk can't
    be silently reversed without organizer review. The middleware lets
    REWS through to /api/swag/unmark (since it's under /api/swag/*), so
    this dependency is the actual organizer-only enforcement.
    """
    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    actor_email = (profile.get("email") or "").strip().lower()
    is_org = False
    if actor_email in SANDBOX_ADMIN_EMAILS:
        is_org = True
    else:
        row = (
            db.query(models.Judge)
            .filter(func.lower(models.Judge.email) == actor_email)
            .filter(models.Judge.is_active.is_(True))
            .first()
        )
        if row and (row.role or "").lower() == "organizer":
            is_org = True
    if not is_org:
        raise HTTPException(status_code=403, detail="Only organizers can undo a swag mark.")

    email = req.email.strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="email is required")
    db.query(models.SwagPickup).filter(models.SwagPickup.email == email).delete(synchronize_session=False)
    db.commit()
    roster = {r["email"]: r for r in _swag_roster(db)}
    if email not in roster:
        raise HTTPException(status_code=404, detail="not found after unmark")
    return SwagPersonOut(**roster[email])


# ===== Swag extras: non-team people who still need a t-shirt =====

@app.get("/api/swag/extras", response_model=list[SwagExtraOut])
def list_swag_extras(db: Session = Depends(get_db)) -> list[SwagExtraOut]:
    """List of non-team people enrolled in the swag roster (judges,
    organisers, support, leadership, HR). Used by the organizer admin
    UI to show what's currently in the table + offer delete buttons."""
    try:
        rows = (
            db.query(models.SwagExtra)
            .order_by(models.SwagExtra.category.asc(), models.SwagExtra.name.asc())
            .all()
        )
    except Exception:
        # Table likely doesn't exist yet -- return empty so the UI can
        # show a friendly 'upload to get started' state.
        return []
    return [
        SwagExtraOut(
            id=r.id,
            email=r.email,
            name=r.name,
            tshirt_size=r.tshirt_size,
            country=r.country,
            category=r.category,
            created_at=r.created_at.isoformat() if r.created_at else None,
        )
        for r in rows
    ]


@app.delete("/api/swag/extras/{extra_id}", response_model=dict)
def delete_swag_extra(
    extra_id: int,
    claims: dict = Depends(require_auth),
    db: Session = Depends(get_db),
) -> dict:
    """Remove a single extras row. Restricted to organizers (same posture
    as swag/unmark -- only organizers can edit the roster, REWS volunteers
    can only mark people collected)."""
    profile = build_profile_payload(claims, {})
    actor_email = (profile.get("email") or "").strip().lower()
    is_org = actor_email in SANDBOX_ADMIN_EMAILS
    if not is_org:
        row = (
            db.query(models.Judge)
            .filter(func.lower(models.Judge.email) == actor_email)
            .filter(models.Judge.is_active.is_(True))
            .first()
        )
        is_org = bool(row and (row.role or "").lower() == "organizer")
    if not is_org:
        raise HTTPException(status_code=403, detail="Only organizers can edit the swag-extras roster.")
    extra = db.query(models.SwagExtra).get(extra_id)
    if not extra:
        raise HTTPException(status_code=404, detail="Swag extra not found")
    db.delete(extra)
    db.commit()
    return {"deleted": True, "id": extra_id}


@app.post("/api/swag/extras/import", response_model=SwagExtraImportResult)
async def import_swag_extras(
    file: UploadFile = File(...),
    claims: dict = Depends(require_auth),
    db: Session = Depends(get_db),
) -> SwagExtraImportResult:
    """Bulk-load non-team people who need a swag kit from an xlsx upload.

    Expected columns (case-insensitive header match, in any order):
      - Name
      - Email
      - T-shirt Size  (or 'Size')
      - Country       (or 'Location')
      - Category      (or 'Role' / 'Type')

    Idempotent on email: re-uploading the same file updates existing rows
    rather than duplicating. Emails already on the member/mentor roster
    are skipped (they get a t-shirt via team membership, no extras row
    needed) and reported back as skipped_existing_roster_count so the
    organizer can sanity-check.
    """
    profile = build_profile_payload(claims, {})
    actor_email = (profile.get("email") or "").strip().lower()
    is_org = actor_email in SANDBOX_ADMIN_EMAILS
    if not is_org:
        row = (
            db.query(models.Judge)
            .filter(func.lower(models.Judge.email) == actor_email)
            .filter(models.Judge.is_active.is_(True))
            .first()
        )
        is_org = bool(row and (row.role or "").lower() == "organizer")
    if not is_org:
        raise HTTPException(status_code=403, detail="Only organizers can import the swag-extras roster.")

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Expected an .xlsx/.xls file")

    from openpyxl import load_workbook

    suffix = os.path.splitext(file.filename)[1] or ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp.flush()
        path = tmp.name

    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active

        # Locate columns by case-insensitive header. Some headers are aliases
        # (Size for T-shirt Size, Location for Country, Role for Category)
        # so the organizer doesn't have to rename their existing spreadsheet.
        headers: dict[str, int] = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            key = str(cell.value or "").strip().lower()
            if key:
                headers[key] = col_idx
        def _col(*names: str) -> int | None:
            for n in names:
                if n.lower() in headers:
                    return headers[n.lower()]
            return None
        name_col = _col("name", "full name")
        email_col = _col("email", "email id", "email ids")
        size_col = _col("t-shirt size", "tshirt size", "size", "shirt size")
        country_col = _col("country", "location", "based out of")
        category_col = _col("category", "role", "type", "group")
        if not name_col or not email_col or not category_col:
            raise HTTPException(
                status_code=400,
                detail=f"Required headers missing. Need at minimum Name, Email, and Category. Found: {sorted(headers.keys())}",
            )

        # Build the lookup of emails already on the team-member / mentor list
        # so we can skip rows that don't need an extras entry.
        team_emails: set[str] = set()
        for team in db.query(models.Team).all():
            if team.mentor_email:
                team_emails.add(team.mentor_email.strip().lower())
            for m in team.members:
                if m.email:
                    team_emails.add(m.email.strip().lower())

        created = 0
        updated = 0
        skipped_existing = 0
        failed: list[dict] = []
        by_category: dict[str, int] = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[name_col - 1] or "").strip()
            email = str(row[email_col - 1] or "").strip().lower()
            if not name and not email:
                continue
            if not email or "@" not in email:
                failed.append({"name": name, "email": email, "error": "missing or malformed email"})
                continue
            category = str(row[category_col - 1] or "").strip()
            if not category:
                failed.append({"name": name, "email": email, "error": "missing category"})
                continue
            tshirt = str(row[size_col - 1] or "").strip().upper() if size_col else None
            tshirt = tshirt or None
            country = str(row[country_col - 1] or "").strip() if country_col else None
            country = country or None

            if email in team_emails:
                skipped_existing += 1
                continue

            existing = (
                db.query(models.SwagExtra)
                .filter(func.lower(models.SwagExtra.email) == email)
                .first()
            )
            if existing:
                changed = False
                if existing.name != name:
                    existing.name = name; changed = True
                if existing.tshirt_size != tshirt:
                    existing.tshirt_size = tshirt; changed = True
                if existing.country != country:
                    existing.country = country; changed = True
                if existing.category != category:
                    existing.category = category; changed = True
                if changed:
                    updated += 1
            else:
                db.add(models.SwagExtra(
                    email=email,
                    name=name or email,
                    tshirt_size=tshirt,
                    country=country,
                    category=category,
                ))
                created += 1
            by_category[category] = by_category.get(category, 0) + 1

        db.commit()
        return SwagExtraImportResult(
            created_count=created,
            updated_count=updated,
            skipped_existing_roster_count=skipped_existing,
            failed=failed,
            by_category=by_category,
        )
    finally:
        try:
            os.unlink(path)
        except Exception:
            pass


@app.get("/api/judges/{judge_id}/assignments", response_model=list[JudgeAssignmentOut])
def get_judge_assignments(
    judge_id: int,
    round: int | None = None,
    db: Session = Depends(get_db),
) -> list[models.JudgeAssignment]:
    q = db.query(models.JudgeAssignment).filter(models.JudgeAssignment.judge_id == judge_id)
    if round is not None:
        q = q.filter(models.JudgeAssignment.round == round)
    return q.all()


@app.post("/api/judges/{judge_id}/assignments", response_model=list[JudgeAssignmentOut])
def set_judge_assignments(
    judge_id: int,
    req: JudgeAssignmentSet,
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
    db: Session = Depends(get_db),
) -> list[models.JudgeAssignment]:
    """Replace the set of teams assigned to this judge for the given round.

    Existing assignments for (judge_id, round) are wiped and replaced with
    the supplied team_ids. Idempotent — sending the same list twice is fine.
    """
    judge = db.query(models.Judge).get(judge_id)
    if not judge:
        raise HTTPException(status_code=404, detail="judge not found")
    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    by_email = (profile.get("email") or "").strip().lower() or None

    db.query(models.JudgeAssignment).filter(
        models.JudgeAssignment.judge_id == judge_id,
        models.JudgeAssignment.round == req.round,
    ).delete(synchronize_session=False)

    new_rows: list[models.JudgeAssignment] = []
    for tid in set(req.team_ids):
        if not db.query(models.Team).get(tid):
            continue  # silently skip team ids that don't exist
        row = models.JudgeAssignment(
            judge_id=judge_id,
            team_id=tid,
            round=req.round,
            assigned_by_email=by_email,
        )
        db.add(row)
        new_rows.append(row)
    db.commit()
    return new_rows


@app.post("/api/judging/scores", response_model=JudgeScoreOut)
def submit_judge_score(req: JudgeScoreSubmit, db: Session = Depends(get_db)) -> models.JudgeScore:
    judge = db.query(models.Judge).get(req.judge_id)
    if not judge:
        raise HTTPException(status_code=404, detail="judge not found")
    team = db.query(models.Team).get(req.team_id)
    if not team:
        raise HTTPException(status_code=404, detail="team not found")
    try:
        record = judging.submit_score(
            db=db,
            judge_id=req.judge_id,
            team_id=req.team_id,
            round_num=req.round,
            scores=req.scores,
            comment=req.comment,
            entered_by_email=req.entered_by_email,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    db.commit()
    return record


@app.get("/api/judging/scores", response_model=list[JudgeScoreOut])
def list_judge_scores(
    round: int | None = None,
    judge_id: int | None = None,
    team_id: int | None = None,
    db: Session = Depends(get_db),
) -> list[models.JudgeScore]:
    q = db.query(models.JudgeScore)
    if round is not None:
        q = q.filter(models.JudgeScore.round == round)
    if judge_id is not None:
        q = q.filter(models.JudgeScore.judge_id == judge_id)
    if team_id is not None:
        q = q.filter(models.JudgeScore.team_id == team_id)
    return q.order_by(models.JudgeScore.submitted_at.desc()).all()


@app.delete("/api/judging/scores", response_model=dict)
def delete_judge_score(
    judge_id: int,
    team_id: int,
    round: int,
    db: Session = Depends(get_db),
) -> dict:
    """Delete a single judge-team-round score row.

    Used by JudgeMode's 'Reset my score' action so a judge can correct
    or rescind a score before it's final. Returns {deleted: bool}.
    """
    score = (
        db.query(models.JudgeScore)
        .filter(
            models.JudgeScore.judge_id == judge_id,
            models.JudgeScore.team_id == team_id,
            models.JudgeScore.round == round,
        )
        .first()
    )
    if not score:
        return {"deleted": False, "message": "No score found for this judge/team/round."}
    db.delete(score)
    db.commit()
    return {"deleted": True, "judge_id": judge_id, "team_id": team_id, "round": round}


@app.get("/api/judging/leaderboard", response_model=LeaderboardOut)
def get_leaderboard(round: int = 1, db: Session = Depends(get_db)) -> LeaderboardOut:
    rows = judging.leaderboard(db, round)
    return LeaderboardOut(round=round, rows=[LeaderboardRow(**r) for r in rows])


# ===== Teams comms + audit log =====

@app.get("/api/comms/mode")
def comms_mode() -> dict:
    return {"mode": comms.GRAPH_MODE}


@app.post("/api/comms/channels", response_model=dict)
def create_channels(
    req: TeamChannelCreateRequest,
    request: Request,
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
    db: Session = Depends(get_db),
) -> dict:
    """Bulk-create Teams channels for the supplied team_ids (or every team
    without one if team_ids is None).

    Uses the real-Graph delegated path same as the per-team button. Each
    iteration calls comms.create_team_channel_with_graph_token; per-team
    errors are collected so one bad team doesn't kill the batch.
    """
    from .db import is_sandbox_request
    sandbox = is_sandbox_request(request)

    graph_token = request.headers.get("x-graph-token", "")
    if not sandbox and not graph_token:
        raise HTTPException(status_code=400, detail="Missing X-Graph-Token header")

    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    organizer_email = (profile.get("email") or "").strip().lower() or req.sent_by_email

    q = db.query(models.Team)
    if req.team_ids is not None:
        q = q.filter(models.Team.id.in_(req.team_ids))
    teams = q.order_by(models.Team.name.asc()).all()

    created: list[dict] = []
    already: list[dict] = []
    failed: list[dict] = []

    import time

    for t in teams:
        if t.has_teams_channel:
            already.append({"team_id": t.id, "team_name": t.name})
            continue
        try:
            result = comms.create_team_channel_with_graph_token(
                db=db, team=t, graph_token=graph_token,
                sent_by_email=organizer_email, sandbox=sandbox,
            )
            db.commit()
            created.append({
                "team_id": t.id,
                "team_name": t.name,
                "channel_id": result.get("channel_id"),
            })
        except comms._GraphChannelError as e:
            db.rollback()
            failed.append({
                "team_id": t.id,
                "team_name": t.name,
                "error": (e.message or "")[:200],
            })
        except Exception as e:
            db.rollback()
            failed.append({
                "team_id": t.id,
                "team_name": t.name,
                "error": f"unexpected: {str(e)[:200]}",
            })

        time.sleep(0.4)  # gentle pacing for Graph throttle

    return {
        "created_count": len(created),
        "already_existing_count": len(already),
        "failed_count": len(failed),
        "created": created,
        "already_existing": already,
        "failed": failed,
        "mode": "graph" if not sandbox else "sandbox",
    }


@app.post("/api/comms/adopt-orphan-channels", response_model=dict)
def adopt_orphan_channels(
    request: Request,
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
    db: Session = Depends(get_db),
) -> dict:
    """Scan the parent Team's channels via Graph and adopt any that match a
    team in our DB which doesn't have a channel recorded.

    Useful when channels were created in an earlier run (e.g. as private
    channels that are no longer tracked) but the Teams channel still
    exists. We don't recreate -- we just point our DB at the existing
    channel id so subsequent posts and the audit trail line up.
    """
    import httpx

    graph_token = request.headers.get("x-graph-token", "")
    if not graph_token:
        raise HTTPException(status_code=400, detail="Missing X-Graph-Token header")
    if not comms.PARENT_TEAM_ID:
        raise HTTPException(status_code=500, detail="GRAPH_PARENT_TEAM_ID not set on the server")

    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    organizer_email = (profile.get("email") or "").strip().lower() or None

    # 1) List every channel in the parent Team. Paginate via @odata.nextLink.
    existing: dict[str, str] = {}  # display_name_lower -> channel_id
    with httpx.Client(
        headers={"Authorization": f"Bearer {graph_token}"},
        timeout=30,
    ) as gc:
        url: str | None = f"{comms.GRAPH}/teams/{comms.PARENT_TEAM_ID}/channels"
        while url:
            r = gc.get(url)
            if r.status_code != 200:
                raise HTTPException(
                    status_code=502,
                    detail=f"Graph error listing parent Team channels ({r.status_code}): {r.text[:800]}",
                )
            data = r.json()
            for ch in data.get("value", []):
                name = (ch.get("displayName") or "").strip()
                cid = ch.get("id")
                if name and cid:
                    existing[name.lower()] = cid
            url = data.get("@odata.nextLink")

    # 2) For each team without a tracked channel, see if a matching channel
    #    already exists in Teams. Match by the same sanitized display name
    #    we would have used at create time.
    teams_without = (
        db.query(models.Team)
        .filter(models.Team.has_teams_channel.is_(False))
        .order_by(models.Team.name.asc())
        .all()
    )

    adopted: list[dict] = []
    not_found: list[dict] = []
    for team in teams_without:
        expected = comms._sanitize_channel_name(f"2026 {team.name}", max_len=50)
        match_id = existing.get(expected.lower())
        if not match_id:
            not_found.append({
                "team_id": team.id,
                "team_name": team.name,
                "expected_name": expected,
            })
            continue
        team.has_teams_channel = True
        team.teams_channel_id = match_id
        team.teams_channel_created_at = datetime.utcnow()
        comms.log(
            db, team.id, kind="teams_channel_create",
            subject=f"Channel adopted: {team.name}",
            body=f"Channel ID: {match_id} · adopted existing channel '{expected}'",
            status="sent",
            sent_by_email=organizer_email,
        )
        adopted.append({
            "team_id": team.id,
            "team_name": team.name,
            "channel_id": match_id,
            "display_name": expected,
        })

    db.commit()
    return {
        "parent_team_channel_count": len(existing),
        "teams_without_channel_in_db": len(teams_without),
        "adopted_count": len(adopted),
        "not_found_count": len(not_found),
        "adopted": adopted,
        "not_found": not_found,
    }


@app.post("/api/comms/check-welcome-mentions", response_model=dict)
def check_welcome_mentions(
    request: Request,
    claims: dict = Depends(require_auth),
    db: Session = Depends(get_db),
) -> dict:
    """For every team in the DB, attempt to resolve mentor + each member's
    email to an Azure AD user id. The ones that DON'T resolve are the
    same people that would have been silently skipped from the welcome
    channel message's @mention list.

    Same root cause covers two organizer questions:
      - Who isn't in the parent Team / can't see the channel? -> usually
        these unresolved emails.
      - Who didn't get an @mention notification when welcome posted? ->
        exactly these unresolved emails.

    Uses User.ReadBasic.All (already granted). Parallelized so 500 lookups
    finish in ~10-15 seconds instead of 90.
    """
    import httpx
    from concurrent.futures import ThreadPoolExecutor

    graph_token = request.headers.get("x-graph-token", "")
    if not graph_token:
        raise HTTPException(status_code=400, detail="Missing X-Graph-Token header")

    teams = db.query(models.Team).order_by(models.Team.name.asc()).all()

    # 1) Collect unique emails with the list of (team, role, name) they
    #    appear under. Dedup by lowercased email so a mentor who covers
    #    multiple teams only gets looked up once.
    email_contexts: dict[str, list[dict]] = {}
    for team in teams:
        if team.mentor_email and team.mentor_name:
            key = team.mentor_email.strip().lower()
            email_contexts.setdefault(key, []).append({
                "team_id": team.id, "team_name": team.name,
                "role": "mentor", "name": team.mentor_name.strip(),
                "email": team.mentor_email.strip(),
            })
        for m in team.members:
            if m.email and m.name:
                key = m.email.strip().lower()
                email_contexts.setdefault(key, []).append({
                    "team_id": team.id, "team_name": team.name,
                    "role": "member", "name": m.name.strip(),
                    "email": m.email.strip(),
                })

    unique_emails = list(email_contexts.keys())
    unresolved: dict[str, int] = {}  # email_lower -> http status

    headers = {"Authorization": f"Bearer {graph_token}"}
    with httpx.Client(headers=headers, timeout=15) as gc:
        def check_one(email: str) -> tuple[str, int]:
            try:
                r = gc.get(f"{comms.GRAPH}/users/{email}")
                return email, r.status_code
            except Exception:
                return email, -1  # network error etc.

        with ThreadPoolExecutor(max_workers=10) as exe:
            for email, status in exe.map(check_one, unique_emails):
                if status != 200:
                    unresolved[email] = status

    # 2) Build per-team report (only teams with at least one unresolved person)
    per_team: dict[int, dict] = {}
    for email, _status in unresolved.items():
        for ctx in email_contexts[email]:
            tid = ctx["team_id"]
            if tid not in per_team:
                per_team[tid] = {
                    "team_id": tid,
                    "team_name": ctx["team_name"],
                    "unresolved": [],
                }
            per_team[tid]["unresolved"].append({
                "role": ctx["role"],
                "name": ctx["name"],
                "email": ctx["email"],
            })

    # Sort each team's unresolved list (mentor first, then alphabetical)
    for v in per_team.values():
        v["unresolved"].sort(key=lambda x: (0 if x["role"] == "mentor" else 1, x["name"].lower()))

    issues_sorted = sorted(per_team.values(), key=lambda t: t["team_name"].lower())

    return {
        "total_teams": len(teams),
        "total_unique_emails": len(unique_emails),
        "resolved_count": len(unique_emails) - len(unresolved),
        "unresolved_email_count": len(unresolved),
        "teams_with_issues_count": len(issues_sorted),
        "teams_with_issues": issues_sorted,
    }


@app.get("/api/comms/welcomed-team-ids", response_model=dict)
def list_welcomed_team_ids(
    claims: dict = Depends(require_auth),
    db: Session = Depends(get_db),
) -> dict:
    """Returns the set of team_ids whose channels have already received the
    RealHack welcome message (detected via comm_log entries with subject
    starting 'Welcome message posted'). The frontend uses this to compute
    'remaining' counts and skip already-done teams in bulk operations.
    """
    rows = (
        db.query(models.CommLog.team_id)
        .filter(models.CommLog.kind == "teams_message")
        .filter(models.CommLog.subject.like("Welcome message posted%"))
        .distinct()
        .all()
    )
    return {"team_ids": sorted({r[0] for r in rows})}


@app.get("/api/comms/qr-posted-team-ids", response_model=dict)
def list_qr_posted_team_ids(
    claims: dict = Depends(require_auth),
    db: Session = Depends(get_db),
) -> dict:
    """Returns the set of team_ids whose channels have already received the
    floor-walk QR-code message (detected via comm_log entries with subject
    starting 'QR-code message posted'). Used by the bulk-QR button to show
    a 'remaining' count instead of always saying 95.
    """
    rows = (
        db.query(models.CommLog.team_id)
        .filter(models.CommLog.kind == "teams_message")
        .filter(models.CommLog.subject.like("QR-code message posted%"))
        .distinct()
        .all()
    )
    return {"team_ids": sorted({r[0] for r in rows})}


@app.get("/api/comms/repo-ready-posted-team-ids", response_model=dict)
def list_repo_ready_posted_team_ids(
    claims: dict = Depends(require_auth),
    db: Session = Depends(get_db),
) -> dict:
    """team_ids whose channels have already received the 'your GitHub repo
    is ready' announcement. Drives the bulk button's 'N remaining' count."""
    rows = (
        db.query(models.CommLog.team_id)
        .filter(models.CommLog.kind == "teams_message")
        .filter(models.CommLog.subject.like("Repo-ready message posted%"))
        .distinct()
        .all()
    )
    return {"team_ids": sorted({r[0] for r in rows})}


@app.post("/api/import/repo-urls", response_model=dict)
async def import_repo_urls(
    file: UploadFile = File(...),
    claims: dict = Depends(require_auth),
    db: Session = Depends(get_db),
) -> dict:
    """Bulk-update Team.repo_url from the DevOps hand-off xlsx.

    Expects the same shape we exported in /api/export/devops-repos.xlsx:
    columns include 'Team Name' and 'Gitrepo url' (case-insensitive header
    match). Reads only the FIRST row of each team's merged block — DevOps
    keeps the URL on the team-name row and leaves member-rows blank for
    those columns, so this is the natural place to find the value.

    Matches by team name (case-insensitive, whitespace-trimmed). Returns
    per-team counts so the user can see which names didn't match and chase
    them up out-of-band rather than guessing fuzzy matches we'd get wrong.
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Expected an .xlsx/.xls file")

    from openpyxl import load_workbook

    suffix = os.path.splitext(file.filename)[1] or ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp.flush()
        path = tmp.name

    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active

        # Find the header row + locate the Team Name + Gitrepo url columns.
        # Be generous on header name variants since DevOps may rename them.
        headers: dict[str, int] = {}
        header_row_idx = 1
        for col_idx, cell in enumerate(ws[1], start=1):
            key = str(cell.value or "").strip().lower()
            if key:
                headers[key] = col_idx
        team_name_col = next(
            (headers[k] for k in ("team name", "team", "team_name") if k in headers),
            None,
        )
        repo_col = next(
            (headers[k] for k in ("gitrepo url", "git repo url", "repo url", "repo_url", "github repo", "github url") if k in headers),
            None,
        )
        if not team_name_col or not repo_col:
            raise HTTPException(
                status_code=400,
                detail=f"Couldn't find 'Team Name' + 'Gitrepo url' columns in the header row. Found: {sorted(headers.keys())}",
            )

        updates: list[dict] = []
        not_found: list[str] = []
        no_repo: list[str] = []
        skipped_existing_same: list[str] = []

        # Iterate data rows. Merged cells: only the top-left of each merge
        # has a value; subsequent rows in the merge return None. That's
        # exactly what we want -- we only need to read the row where Team
        # Name is filled in.
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            team_name = str(row[team_name_col - 1] or "").strip()
            if not team_name:
                continue
            repo_url = str(row[repo_col - 1] or "").strip()
            if not repo_url or repo_url.lower() in ("url", "tbd", "n/a", "na"):
                no_repo.append(team_name)
                continue
            # Case-insensitive name match. Don't fuzzy-match -- safer to
            # report a miss and let the organizer correct the spreadsheet.
            team = (
                db.query(models.Team)
                .filter(func.lower(models.Team.name) == team_name.lower())
                .first()
            )
            if not team:
                not_found.append(team_name)
                continue
            if (team.repo_url or "").strip() == repo_url:
                skipped_existing_same.append(team_name)
                continue
            team.repo_url = repo_url
            updates.append({"team_id": team.id, "team_name": team.name, "repo_url": repo_url})

        db.commit()

        return {
            "updated_count": len(updates),
            "unchanged_count": len(skipped_existing_same),
            "not_found_count": len(not_found),
            "no_repo_url_count": len(no_repo),
            "updated": updates[:50],  # cap for UI; full list visible in DB
            "not_found": not_found,
            "no_repo_url": no_repo,
        }
    finally:
        try:
            os.unlink(path)
        except Exception:
            pass


@app.post("/api/comms/teams/{team_id}/adopt-channel-by-link", response_model=dict)
def adopt_channel_by_link(
    team_id: int,
    req: AdoptChannelByLinkRequest,
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
    db: Session = Depends(get_db),
) -> dict:
    """Adopt an existing Microsoft Teams channel by pasting its share link.

    Used when a channel exists in Teams but isn't tracked in our DB AND
    the auto-discovery via Graph isn't an option (would need a Channel
    or Group read scope that isn't admin-consented in our tenant).

    Accepts either:
      - A 'Get link to channel' URL from Teams, e.g.
          https://teams.microsoft.com/l/channel/19%3Axxx%40thread.tacv2/...
      - A raw channel id like '19:xxx@thread.tacv2'

    Updates the team's DB row to point at that channel; no Graph call.
    """
    import re
    import urllib.parse as _urlparse

    team = db.query(models.Team).get(team_id)
    if not team:
        raise HTTPException(status_code=404, detail="team not found")

    raw = (req.teams_channel_link or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="teams_channel_link is required")

    # Extract channel id from the URL pattern .../l/channel/<ENCODED_ID>/...
    m = re.search(r"/l/channel/([^/?#]+)", raw)
    if m:
        channel_id = _urlparse.unquote(m.group(1))
    elif raw.startswith("19:") and "@thread" in raw:
        # Raw channel id pasted directly
        channel_id = raw
    else:
        raise HTTPException(
            status_code=400,
            detail=(
                "Could not parse channel id. Paste either a 'Get link to channel' URL "
                "from Teams (right-click channel -> Get link), or the raw channel id "
                "starting with '19:' and ending with '@thread.tacv2'."
            ),
        )

    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    organizer_email = (profile.get("email") or "").strip().lower() or None

    team.has_teams_channel = True
    team.teams_channel_id = channel_id
    team.teams_channel_created_at = datetime.utcnow()
    comms.log(
        db, team.id, kind="teams_channel_create",
        subject=f"Channel adopted (manual link): {team.name}",
        body=f"Channel ID: {channel_id} · adopted via paste-link",
        status="sent",
        sent_by_email=organizer_email,
    )
    db.commit()
    return {
        "team_id": team.id,
        "team_name": team.name,
        "channel_id": channel_id,
        "status": "adopted",
    }


@app.post("/api/comms/teams/{team_id}/reset-channel", response_model=dict)
def reset_team_channel_state(
    team_id: int,
    claims: dict = Depends(require_auth),
    db: Session = Depends(get_db),
) -> dict:
    """Reset a team's channel state in our DB (clear has_teams_channel,
    teams_channel_id, and the welcome-posted audit log entry). Use after
    manually deleting a team's Teams channel in the Teams UI so a fresh
    channel can be created and welcome reposted.

    Does NOT call Graph — this only cleans up DB state to match a manual
    Teams-UI deletion the organizer already performed.
    """
    team = db.query(models.Team).get(team_id)
    if not team:
        raise HTTPException(status_code=404, detail="team not found")

    prev_channel_id = team.teams_channel_id
    team.has_teams_channel = False
    team.teams_channel_id = None
    team.teams_channel_created_at = None

    # Remove the 'Welcome message posted' audit entries so the bulk
    # welcome button picks this team up again on the next click.
    deleted = (
        db.query(models.CommLog)
        .filter(models.CommLog.team_id == team.id)
        .filter(models.CommLog.kind == "teams_message")
        .filter(models.CommLog.subject.like("Welcome message posted%"))
        .delete(synchronize_session=False)
    )

    db.commit()
    return {
        "team_id": team.id,
        "team_name": team.name,
        "previous_channel_id": prev_channel_id,
        "welcome_log_entries_removed": deleted,
        "status": "reset",
    }


@app.post("/api/comms/teams/{team_id}/create-channel", response_model=dict)
def create_one_team_channel(
    team_id: int,
    request: Request,
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
    db: Session = Depends(get_db),
) -> dict:
    """Create a single Teams channel for the given team using a delegated
    Graph token supplied by the browser. Per-team button on the dashboard.

    The Graph token comes in the `X-Graph-Token` header — the frontend
    acquires it via MSAL.acquireTokenSilent({scopes: [Channel.Create, ...]}).
    In sandbox/Test Mode, the Graph call is skipped and a mock entry is
    written instead so the UI flow can be tested without making real channels.
    """
    team = db.query(models.Team).get(team_id)
    if not team:
        raise HTTPException(status_code=404, detail="team not found")
    if team.has_teams_channel:
        raise HTTPException(status_code=409, detail="Team already has a channel")

    from .db import is_sandbox_request
    sandbox = is_sandbox_request(request)

    graph_token = request.headers.get("x-graph-token", "")
    if not sandbox and not graph_token:
        raise HTTPException(status_code=400, detail="Missing X-Graph-Token header")

    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    organizer_email = (profile.get("email") or "").strip().lower() or None

    try:
        result = comms.create_team_channel_with_graph_token(
            db=db,
            team=team,
            graph_token=graph_token,
            sent_by_email=organizer_email,
            sandbox=sandbox,
        )
    except comms._GraphChannelError as e:
        db.rollback()
        raise HTTPException(status_code=e.status, detail=e.message)
    db.commit()
    return result


@app.post("/api/comms/teams/{team_id}/message", response_model=dict)
def post_team_message(team_id: int, req: TeamMessageRequest, db: Session = Depends(get_db)) -> dict:
    team = db.query(models.Team).get(team_id)
    if not team:
        raise HTTPException(status_code=404, detail="team not found")
    result = comms.post_team_message(db, team, req.message, sent_by_email=req.sent_by_email)
    db.commit()
    return result


@app.post("/api/comms/teams/{team_id}/post-channel-qr", response_model=dict)
def post_channel_qr_for_team(
    team_id: int,
    request: Request,
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
    db: Session = Depends(get_db),
) -> dict:
    """Post the floor-walk QR-code message to ONE team's channel."""
    team = db.query(models.Team).get(team_id)
    if not team:
        raise HTTPException(status_code=404, detail="team not found")
    if not team.has_teams_channel or not team.teams_channel_id:
        raise HTTPException(status_code=409, detail="Team has no Teams channel yet")
    if str(team.teams_channel_id).startswith(("sandbox-", "mock-", "dryrun-")):
        raise HTTPException(status_code=400, detail="Cannot post to a mock channel")

    graph_token = request.headers.get("x-graph-token", "")
    if not graph_token:
        raise HTTPException(status_code=400, detail="Missing X-Graph-Token header")

    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    organizer_email = (profile.get("email") or "").strip().lower() or None

    public_base = str(request.base_url).rstrip("/")
    # request.base_url is the backend URL (e.g. http://127.0.0.1:8001). The
    # QR needs to point at the public frontend origin. Prefer the explicit
    # PUBLIC_BASE_URL env if set, otherwise fall back to the request Origin
    # header (sent by browser), otherwise base_url.
    public_base = (
        os.environ.get("PUBLIC_BASE_URL")
        or request.headers.get("origin")
        or public_base
    )

    try:
        result = comms.post_channel_qr_message_with_graph_token(
            db=db, team=team, graph_token=graph_token,
            public_base_url=public_base,
            sent_by_email=organizer_email,
        )
    except comms._GraphChannelError as e:
        db.rollback()
        raise HTTPException(status_code=e.status, detail=e.message)
    db.commit()
    return result


@app.post("/api/comms/teams/{team_id}/post-channel-repo-ready", response_model=dict)
def post_channel_repo_ready_for_team(
    team_id: int,
    request: Request,
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
    db: Session = Depends(get_db),
) -> dict:
    """Post the 'your GitHub repo is ready' announcement to ONE team's channel.
    Used to sanity-check the message format on a single team before firing
    the bulk button at all 95. No idempotency check on the per-team route --
    organizers may need to re-test after a template tweak."""
    team = db.query(models.Team).get(team_id)
    if not team:
        raise HTTPException(status_code=404, detail="team not found")
    if not team.has_teams_channel or not team.teams_channel_id:
        raise HTTPException(status_code=409, detail="Team has no Teams channel yet")
    if str(team.teams_channel_id).startswith(("sandbox-", "mock-", "dryrun-")):
        raise HTTPException(status_code=400, detail="Cannot post to a mock channel")
    if not (team.repo_url or "").strip():
        raise HTTPException(status_code=409, detail="Team has no repo_url on file — import the DevOps xlsx first")

    graph_token = request.headers.get("x-graph-token", "")
    if not graph_token:
        raise HTTPException(status_code=400, detail="Missing X-Graph-Token header")

    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    organizer_email = (profile.get("email") or "").strip().lower() or None

    try:
        result = comms.post_channel_repo_ready_with_graph_token(
            db=db, team=team, graph_token=graph_token,
            sent_by_email=organizer_email,
        )
    except comms._GraphChannelError as e:
        db.rollback()
        raise HTTPException(status_code=e.status, detail=e.message)
    db.commit()
    return result


@app.post("/api/comms/teams/post-channel-qr-all", response_model=dict)
def post_channel_qr_to_all(
    request: Request,
    force: bool = False,
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
    db: Session = Depends(get_db),
) -> dict:
    """Bulk: post the QR-code message to every team that has a real channel.
    Idempotent by default — skips teams that already received a 'QR-code
    message posted' log entry. Pass `?force=true` to re-post to every team
    regardless (used when the message template itself changes — e.g. we
    added the floor-walk seat-info call to action and need everyone to
    see the updated text). Per-team failures collected; batch continues."""
    graph_token = request.headers.get("x-graph-token", "")
    if not graph_token:
        raise HTTPException(status_code=400, detail="Missing X-Graph-Token header")

    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    organizer_email = (profile.get("email") or "").strip().lower() or None

    public_base = (
        os.environ.get("PUBLIC_BASE_URL")
        or request.headers.get("origin")
        or str(request.base_url).rstrip("/")
    )

    all_teams = (
        db.query(models.Team)
        .filter(models.Team.has_teams_channel.is_(True))
        .filter(models.Team.teams_channel_id.isnot(None))
        .order_by(models.Team.name.asc())
        .all()
    )

    posted: list[dict] = []
    skipped_already: list[dict] = []
    skipped_no_real_channel: list[dict] = []
    failed: list[dict] = []

    import time
    for team in all_teams:
        if str(team.teams_channel_id).startswith(("sandbox-", "mock-", "dryrun-")):
            skipped_no_real_channel.append({"team_id": team.id, "team_name": team.name})
            continue
        if not force:
            already = (
                db.query(models.CommLog)
                .filter(models.CommLog.team_id == team.id)
                .filter(models.CommLog.kind == "teams_message")
                .filter(models.CommLog.subject.like("QR-code message posted%"))
                .first()
            )
            if already:
                skipped_already.append({"team_id": team.id, "team_name": team.name})
                continue
        try:
            r = comms.post_channel_qr_message_with_graph_token(
                db=db, team=team, graph_token=graph_token,
                public_base_url=public_base,
                sent_by_email=organizer_email,
            )
            db.commit()
            posted.append({"team_id": team.id, "team_name": team.name})
        except comms._GraphChannelError as e:
            db.rollback()
            failed.append({"team_id": team.id, "team_name": team.name, "error": (e.message or "")[:200]})
        except Exception as e:
            db.rollback()
            failed.append({"team_id": team.id, "team_name": team.name, "error": f"unexpected: {str(e)[:200]}"})
        time.sleep(0.5)

    return {
        "total_teams_with_channels": len(all_teams),
        "posted_count": len(posted),
        "skipped_already_posted_count": len(skipped_already),
        "skipped_no_real_channel_count": len(skipped_no_real_channel),
        "failed_count": len(failed),
        "posted": posted,
        "skipped_already_posted": skipped_already,
        "skipped_no_real_channel": skipped_no_real_channel,
        "failed": failed,
    }


@app.post("/api/comms/teams/post-channel-repo-ready-all", response_model=dict)
def post_channel_repo_ready_to_all(
    request: Request,
    force: bool = False,
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
    db: Session = Depends(get_db),
) -> dict:
    """Bulk: announce the provisioned GitHub repo in every team's channel.

    Skips teams that:
      - have no Teams channel yet, or
      - have a mock/sandbox channel, or
      - have no repo_url on file (import the DevOps xlsx first), or
      - already received a 'Repo-ready message posted' log entry (unless
        ?force=true to override the idempotency check, used when the
        message template itself changes and we need everyone to see the
        updated text).

    Per-team failures are collected; the batch always continues.
    """
    graph_token = request.headers.get("x-graph-token", "")
    if not graph_token:
        raise HTTPException(status_code=400, detail="Missing X-Graph-Token header")

    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    organizer_email = (profile.get("email") or "").strip().lower() or None

    all_teams = (
        db.query(models.Team)
        .filter(models.Team.has_teams_channel.is_(True))
        .filter(models.Team.teams_channel_id.isnot(None))
        .order_by(models.Team.name.asc())
        .all()
    )

    posted: list[dict] = []
    skipped_already: list[dict] = []
    skipped_no_real_channel: list[dict] = []
    skipped_no_repo: list[dict] = []
    failed: list[dict] = []

    import time
    for team in all_teams:
        if str(team.teams_channel_id).startswith(("sandbox-", "mock-", "dryrun-")):
            skipped_no_real_channel.append({"team_id": team.id, "team_name": team.name})
            continue
        if not (team.repo_url or "").strip():
            skipped_no_repo.append({"team_id": team.id, "team_name": team.name})
            continue
        if not force:
            already = (
                db.query(models.CommLog)
                .filter(models.CommLog.team_id == team.id)
                .filter(models.CommLog.kind == "teams_message")
                .filter(models.CommLog.subject.like("Repo-ready message posted%"))
                .first()
            )
            if already:
                skipped_already.append({"team_id": team.id, "team_name": team.name})
                continue
        try:
            comms.post_channel_repo_ready_with_graph_token(
                db=db, team=team, graph_token=graph_token,
                sent_by_email=organizer_email,
            )
            db.commit()
            posted.append({"team_id": team.id, "team_name": team.name})
        except comms._GraphChannelError as e:
            db.rollback()
            failed.append({"team_id": team.id, "team_name": team.name, "error": (e.message or "")[:200]})
        except Exception as e:
            db.rollback()
            failed.append({"team_id": team.id, "team_name": team.name, "error": f"unexpected: {str(e)[:200]}"})
        time.sleep(0.5)

    return {
        "total_teams_with_channels": len(all_teams),
        "posted_count": len(posted),
        "skipped_already_posted_count": len(skipped_already),
        "skipped_no_real_channel_count": len(skipped_no_real_channel),
        "skipped_no_repo_url_count": len(skipped_no_repo),
        "failed_count": len(failed),
        "posted": posted,
        "skipped_already_posted": skipped_already,
        "skipped_no_real_channel": skipped_no_real_channel,
        "skipped_no_repo_url": skipped_no_repo,
        "failed": failed,
    }


@app.post("/api/comms/teams/post-channel-welcome-all", response_model=dict)
def post_channel_welcome_to_all(
    request: Request,
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
    db: Session = Depends(get_db),
) -> dict:
    """Post the RealHack 2026 welcome message to every team that has a real
    Teams channel. Idempotent: skips any team that already has a 'Welcome
    message posted' entry in the comm log. Continues on per-team failures
    so one bad team doesn't sink the whole batch.

    Synchronous — takes ~2-3 min for 95 teams. The frontend keeps the
    busy state until this returns the summary.
    """
    graph_token = request.headers.get("x-graph-token", "")
    if not graph_token:
        raise HTTPException(status_code=400, detail="Missing X-Graph-Token header")

    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    organizer_email = (profile.get("email") or "").strip().lower() or None

    # All teams that claim to have a channel (filter mock/sandbox later)
    all_teams = (
        db.query(models.Team)
        .filter(models.Team.has_teams_channel.is_(True))
        .filter(models.Team.teams_channel_id.isnot(None))
        .order_by(models.Team.name.asc())
        .all()
    )

    posted: list[dict] = []
    skipped_already: list[dict] = []
    skipped_no_real_channel: list[dict] = []
    failed: list[dict] = []

    import time

    for team in all_teams:
        # Skip mock / sandbox channel ids
        if str(team.teams_channel_id).startswith(("sandbox-", "mock-", "dryrun-")):
            skipped_no_real_channel.append({"team_id": team.id, "team_name": team.name})
            continue

        # Skip if a welcome was already posted (idempotency check via comm log)
        already = (
            db.query(models.CommLog)
            .filter(models.CommLog.team_id == team.id)
            .filter(models.CommLog.kind == "teams_message")
            .filter(models.CommLog.subject.like("Welcome message posted%"))
            .first()
        )
        if already:
            skipped_already.append({"team_id": team.id, "team_name": team.name})
            continue

        try:
            r = comms.post_channel_welcome_with_graph_token(
                db=db, team=team, graph_token=graph_token, sent_by_email=organizer_email,
            )
            db.commit()
            posted.append({
                "team_id": team.id,
                "team_name": team.name,
                "mentions": r["mentions_count"],
            })
        except comms._GraphChannelError as e:
            db.rollback()
            failed.append({
                "team_id": team.id,
                "team_name": team.name,
                "error": (e.message or "")[:200],
            })
        except Exception as e:
            db.rollback()
            failed.append({
                "team_id": team.id,
                "team_name": team.name,
                "error": f"unexpected: {str(e)[:200]}",
            })

        # Light pacing to stay under Graph throttle limits.
        time.sleep(0.5)

    return {
        "total_teams_with_channels": len(all_teams),
        "posted_count": len(posted),
        "skipped_already_posted_count": len(skipped_already),
        "skipped_no_real_channel_count": len(skipped_no_real_channel),
        "failed_count": len(failed),
        "posted": posted,
        "skipped_already_posted": skipped_already,
        "skipped_no_real_channel": skipped_no_real_channel,
        "failed": failed,
    }


@app.post("/api/comms/teams/{team_id}/post-channel-welcome", response_model=dict)
def post_channel_welcome(
    team_id: int,
    request: Request,
    claims: dict = Depends(require_auth),
    creds: HTTPAuthorizationCredentials = Security(_bearer_scheme),
    db: Session = Depends(get_db),
) -> dict:
    """Post the RealHack 2026 welcome message to a team's Teams channel,
    with @mentions for the mentor and all members.

    Uses the same X-Graph-Token header pattern as the per-team channel
    create. Requires ChannelMessage.Send delegated scope (already
    consented at the tenant level).
    """
    team = db.query(models.Team).get(team_id)
    if not team:
        raise HTTPException(status_code=404, detail="team not found")
    if not team.has_teams_channel or not team.teams_channel_id:
        raise HTTPException(status_code=409, detail="Team has no Teams channel yet")
    if str(team.teams_channel_id).startswith(("sandbox-", "mock-", "dryrun-")):
        raise HTTPException(status_code=400, detail="Cannot post real messages to a mock channel")

    graph_token = request.headers.get("x-graph-token", "")
    if not graph_token:
        raise HTTPException(status_code=400, detail="Missing X-Graph-Token header")

    profile = build_profile_payload(claims, fetch_profile(creds.credentials) if creds else {})
    organizer_email = (profile.get("email") or "").strip().lower() or None

    try:
        result = comms.post_channel_welcome_with_graph_token(
            db=db, team=team, graph_token=graph_token, sent_by_email=organizer_email,
        )
    except comms._GraphChannelError as e:
        db.rollback()
        raise HTTPException(status_code=e.status, detail=e.message)
    db.commit()
    return result


@app.post("/api/comms/broadcast", response_model=dict)
def broadcast_message(req: BroadcastRequest, db: Session = Depends(get_db)) -> dict:
    result = comms.broadcast(db, req.message, team_ids=req.team_ids, sent_by_email=req.sent_by_email)
    db.commit()
    return result


@app.get("/api/comms/log", response_model=list[CommLogOut])
def list_comm_log(
    team_id: int | None = None,
    kind: str | None = None,
    limit: int = 200,
    db: Session = Depends(get_db),
) -> list[models.CommLog]:
    q = db.query(models.CommLog)
    if team_id is not None:
        q = q.filter(models.CommLog.team_id == team_id)
    if kind is not None:
        q = q.filter(models.CommLog.kind == kind)
    return q.order_by(models.CommLog.sent_at.desc()).limit(limit).all()


@app.post("/api/comms/log", response_model=CommLogOut)
def append_comm_log(req: CommLogCreateRequest, db: Session = Depends(get_db)) -> models.CommLog:
    entry = comms.log(
        db,
        team_id=req.team_id,
        kind=req.kind,
        template_id=req.template_id,
        subject=req.subject,
        body=req.body,
        recipients=req.recipients,
        status=req.status,
        sent_by_email=req.sent_by_email,
    )
    db.commit()
    return entry


@app.get("/api/comms/duplicate-check")
def duplicate_check(
    team_id: int,
    kind: str = "email",
    template_id: str | None = None,
    hours: int = 24,
    db: Session = Depends(get_db),
) -> dict:
    entry = comms.recent_duplicate(db, team_id, kind, template_id, hours)
    if not entry:
        return {"duplicate": False}
    return {
        "duplicate": True,
        "last_sent_at": entry.sent_at.isoformat(),
        "last_sent_by_email": entry.sent_by_email,
        "kind": entry.kind,
        "template_id": entry.template_id,
    }


@app.post("/api/teams/{team_id}/check-repo", response_model=RepoCheckOut)
def check_repo(team_id: int, db: Session = Depends(get_db)) -> RepoCheckOut:
    team = db.query(models.Team).get(team_id)
    if not team:
        raise HTTPException(status_code=404, detail="team not found")
    result = comms.check_repo_readiness(team)
    db.commit()
    return RepoCheckOut(**result)


# ===== Backups =====

@app.get("/api/backup/status")
def backup_status() -> dict:
    return backup_service.get_status()


@app.get("/api/backup/list")
def backup_list() -> list[dict]:
    return backup_service.list_backups()


@app.post("/api/backup/now")
def backup_now() -> dict:
    return backup_service.do_backup(reason="manual")


@app.get("/api/backup/download/{filename}")
def backup_download(filename: str):
    from pathlib import Path
    if not filename.startswith("realhack_pilot_") or not filename.endswith(".db") or "/" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="invalid filename")
    path = Path("./backups") / filename
    if not path.exists():
        raise HTTPException(status_code=404, detail="backup not found")
    return FileResponse(str(path), media_type="application/x-sqlite3", filename=filename)


@app.get("/api/backup/download-latest")
def backup_download_latest():
    backups = backup_service.list_backups()
    if not backups:
        raise HTTPException(status_code=404, detail="no backups yet")
    return backup_download(backups[0]["filename"])


@app.post("/api/backup/restore/{filename}")
def backup_restore(filename: str) -> dict:
    if not filename.startswith("realhack_pilot_") or not filename.endswith(".db") or "/" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="invalid filename")
    return backup_service.restore(filename, make_pre_restore_backup=True)


@app.patch("/api/teams/{team_id}/readiness", response_model=TeamOut)
def update_readiness(team_id: int, req: ReadinessFlagsRequest, db: Session = Depends(get_db)) -> models.Team:
    team = db.query(models.Team).get(team_id)
    if not team:
        raise HTTPException(status_code=404, detail="team not found")
    if req.presentation_uploaded is not None:
        team.presentation_uploaded = req.presentation_uploaded
    if req.repo_url is not None:
        team.repo_url = req.repo_url or None
    if req.has_teams_channel is not None:
        team.has_teams_channel = req.has_teams_channel
    db.commit()
    db.refresh(team)
    return team


# ===== Team / member edit (post-registration change requests) =====
# Organizers use these to apply approved change requests against teams
# whose MS Forms entry has already been finalized. Every change writes an
# audit row via comms.log() with kind="team_edit" so we have a permanent
# record of who edited what, when, and why.

_EDITABLE_TEAM_FIELDS = (
    "name", "mentor_name", "mentor_email", "mentor_location", "mentor_tshirt_size",
    "mentor_address",
    "idea", "tools", "approach", "viability", "business_value", "repo_url",
)


def _signed_in_email(claims: dict) -> str | None:
    return claims.get("preferred_username") or claims.get("upn") or claims.get("email")


def _audit_team_edit(
    db: Session, *, team_id: int, summary: str, reason: str | None, editor_email: str | None,
) -> None:
    body = summary if not reason else f"{summary}\nReason: {reason}"
    comms.log(
        db, team_id=team_id, kind="team_edit",
        subject="Team data edited via dashboard",
        body=body,
        status="sent",
        sent_by_email=editor_email,
    )


@app.get("/api/export-msforms.xlsx")
def export_msforms_xlsx(db: Session = Depends(get_db)) -> StreamingResponse:
    """Export current team + member state as an .xlsx whose column layout
    matches what the MS Forms importer reads. Use this BEFORE re-uploading
    a fresh Excel so manual additions / edits don't get wiped — workflow is:

        1. Download this file (current state, MS-Forms-compatible).
        2. Open in Excel, make edits (add late teams, fix typos, etc).
        3. Re-upload via the dashboard's 'Upload registrations' card.

    Format: one row per team. Member slots 1-5; mentor + member t-shirt
    addresses use the 'opted for US or PH as location[N]' columns the
    importer expects.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Registrations"

    # Column layout mirrors EXACTLY the original MS Forms registration export
    # (RealHack 2026 form). Quirks preserved on purpose:
    #   * 'Team Name2' (literal, with the trailing 2 — Forms duplicate-field artefact)
    #   * 'Member1 Name' / 'Member 1 Email' inconsistent spacing
    #   * Non-breaking-space (U+00A0, '\xa0') in t-shirt/location columns
    #   * 'Member4\xa0Location' specifically lacks the space between
    #     'Member' and '4' that other members have.
    # The importer's fuzzy column matcher handles any of these spellings,
    # but matching the source exactly makes diffs on the file readable.
    NBSP = "\xa0"
    headers = [
        "ID", "Start time", "Completion time", "Email", "Name", "Last modified time",
        "Team Name2",
        "Idea/Problem statement", "Tech stack",
        "Approach towards the solution", "Viability of the Solution", "Business value",
        "Mentor Name", "Mentor Email", "Mentor T-shirt Size", "Mentor Location",
        "Enter your mailing address if you opted for US or PH as location",  # mentor address
    ]
    # Members 1-5 — 5 columns each, preserving the exact label idiosyncrasies.
    member_headers = [
        # (name, email, t-shirt, location)
        ("Member1 Name",  "Member 1 Email",  f"Member 1{NBSP}T-shirt Size",  f"Member 1{NBSP}Location"),
        ("Member 2{NBSP}Name".format(NBSP=NBSP), "Member 2 Email", f"Member 2{NBSP}T-shirt Size", f"Member 2{NBSP}Location"),
        ("Member 3{NBSP}Name".format(NBSP=NBSP), "Member 3 Email", f"Member 3{NBSP}T-shirt Size", f"Member 3{NBSP}Location"),
        ("Member 4{NBSP}Name".format(NBSP=NBSP), "Member 4 Email", f"Member 4{NBSP}T-shirt Size", f"Member4{NBSP}Location"),  # Forms typo: no space before 4
        ("Member 5{NBSP}Name".format(NBSP=NBSP), "Member 5 Email", f"Member 5{NBSP}T-shirt Size", f"Member 5{NBSP}Location"),
    ]
    for n, (nm, em, ts, loc) in enumerate(member_headers, start=1):
        headers.extend([nm, em, ts, loc])
        # Address slot N+1 (MS Forms uses '...location2' for member 1, '...location3' for member 2, etc.)
        headers.append(f"Enter your mailing address if you opted for US or PH as location{n + 1}")
    ws.append(headers)

    # Style header row — bold + light fill so it reads as a header in Excel.
    bold = Font(bold=True)
    fill = PatternFill(start_color="FFE3E8F0", end_color="FFE3E8F0", fill_type="solid")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold
        cell.fill = fill

    teams = db.query(models.Team).order_by(models.Team.id.asc()).all()
    for t in teams:
        # First 6 system columns (ID, Start time, Completion time, Email,
        # Name, Last modified time) are blank/synthetic for fresh re-imports.
        submitter_email = t.mentor_email or (t.members[0].email if t.members else "")
        completion_iso = t.submitted_at.isoformat() if t.submitted_at else ""
        row: list = [
            t.external_id or t.id,
            "",                # Start time (not preserved)
            completion_iso,    # Completion time
            submitter_email,   # Email — best-effort: submitter is usually the mentor
            t.mentor_name or "",  # Name (whoever filled in the form)
            "",                # Last modified time
            t.name or "",      # Team Name2
            t.idea or "",
            t.tools or "",
            t.approach or "",
            t.viability or "",
            t.business_value or "",
            t.mentor_name or "",
            t.mentor_email or "",
            t.mentor_tshirt_size or "",
            t.mentor_location or "",
            t.mentor_address or "",
        ]
        # 5 member slots: name, email, t-shirt, location, address
        sorted_members = sorted(t.members, key=lambda m: m.position or 0)
        for n in range(1, 6):
            m = sorted_members[n - 1] if n - 1 < len(sorted_members) else None
            row.extend([
                (m.name if m else "") or "",
                (m.email if m else "") or "",
                (m.tshirt_size if m else "") or "",
                (m.location if m else "") or "",
                (m.address if m else "") or "",
            ])
        ws.append(row)

    # Reasonable column widths so the file is readable in Excel.
    width_hints = {
        "Team Name2": 24, "Mentor Name": 24, "Mentor Email": 30, "Email": 28,
        "Idea/Problem statement": 60, "Tech stack": 40,
        "Approach towards the solution": 50,
        "Viability of the Solution": 50, "Business value": 50,
    }
    for col_idx, header in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width_hints.get(header, 22)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"realhack_2026_registrations_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/teams", response_model=TeamOut)
def create_team_manual(
    req: TeamCreate,
    db: Session = Depends(get_db),
    claims: dict = Depends(require_auth),
) -> models.Team:
    """Manually add a new team to the dashboard (without going through the
    MS Forms Excel import). Used for late registrations, special cases, or
    teams that came in by other channels. Triggers the screener at the end
    so the new team gets a completeness score + flags like any imported team.
    """
    name = (req.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="team name is required")

    # Soft duplicate-check: case-insensitive name match
    existing = db.query(models.Team).filter(models.Team.name.ilike(name)).first()
    if existing:
        raise HTTPException(status_code=409, detail=f"A team named '{name}' already exists")

    team = models.Team(
        name=name,
        mentor_name=(req.mentor_name or "").strip() or None,
        mentor_email=(req.mentor_email or "").strip().lower() or None,
        mentor_location=(req.mentor_location or "").strip() or None,
        mentor_tshirt_size=(req.mentor_tshirt_size or "").strip().upper() or None,
        mentor_address=(req.mentor_address or "").strip() or None,
        idea=(req.idea or "").strip() or None,
        tools=(req.tools or "").strip() or None,
        approach=(req.approach or "").strip() or None,
        viability=(req.viability or "").strip() or None,
        business_value=(req.business_value or "").strip() or None,
        repo_url=(req.repo_url or "").strip() or None,
        raw={},  # no MS Forms row to preserve
    )
    db.add(team)
    db.flush()  # populate team.id before adding members

    for idx, m in enumerate(req.members or [], start=1):
        m_name = (m.name or "").strip()
        if not m_name:
            continue
        db.add(models.Member(
            team_id=team.id,
            name=m_name,
            email=(m.email or "").strip().lower() or None,
            location=(m.location or "").strip() or None,
            tshirt_size=(m.tshirt_size or "").strip().upper() or None,
            address=(m.address or "").strip() or None,
            position=idx,
        ))

    db.flush()

    # Run the screener so completeness + flags get computed before the team
    # appears on the dashboard.
    try:
        screener.screen_all(db)
    except Exception:
        # Don't block creation if the screener hiccups — flags can be recomputed later.
        pass

    db.commit()
    db.refresh(team)
    _audit_team_edit(
        db, team_id=team.id,
        summary=f"Created team '{team.name}' manually with {len(team.members)} member(s)",
        reason=req.edit_reason,
        editor_email=_signed_in_email(claims),
    )
    db.commit()
    return team


@app.patch("/api/teams/{team_id}", response_model=TeamOut)
def patch_team(
    team_id: int,
    req: TeamPatch,
    db: Session = Depends(get_db),
    claims: dict = Depends(require_auth),
) -> models.Team:
    team = db.query(models.Team).get(team_id)
    if not team:
        raise HTTPException(status_code=404, detail="team not found")

    changes: list[str] = []
    for field in _EDITABLE_TEAM_FIELDS:
        new_val = getattr(req, field)
        if new_val is None:
            continue
        old_val = getattr(team, field)
        # Normalise empty string → None for nullable text fields (everything
        # except name, which is required).
        normalised = new_val if field == "name" else (new_val or None)
        if normalised != old_val:
            old_display = old_val if old_val is not None else "(empty)"
            new_display = normalised if normalised is not None else "(empty)"
            changes.append(f"{field}: {old_display!r} -> {new_display!r}")
            setattr(team, field, normalised)

    if not changes:
        # No-op patch: nothing changed, don't write an audit row.
        return team

    db.commit()
    db.refresh(team)
    _audit_team_edit(
        db, team_id=team.id,
        summary="Updated fields: " + "; ".join(changes),
        reason=req.edit_reason,
        editor_email=_signed_in_email(claims),
    )
    db.commit()
    return team


@app.post("/api/teams/{team_id}/members", response_model=MemberOut)
def add_member(
    team_id: int,
    req: MemberCreate,
    db: Session = Depends(get_db),
    claims: dict = Depends(require_auth),
) -> models.Member:
    team = db.query(models.Team).get(team_id)
    if not team:
        raise HTTPException(status_code=404, detail="team not found")
    if not (req.name or "").strip():
        raise HTTPException(status_code=400, detail="name is required")

    next_pos = max((m.position for m in team.members), default=-1) + 1
    member = models.Member(
        team_id=team.id,
        name=req.name.strip(),
        email=(req.email or "").strip() or None,
        location=(req.location or "").strip() or None,
        tshirt_size=(req.tshirt_size or "").strip() or None,
        address=(req.address or "").strip() or None,
        position=next_pos,
    )
    db.add(member)
    db.commit()
    db.refresh(member)
    _audit_team_edit(
        db, team_id=team.id,
        summary=f"Added member: {member.name} <{member.email or 'no-email'}>",
        reason=req.edit_reason,
        editor_email=_signed_in_email(claims),
    )
    db.commit()
    return member


@app.patch("/api/members/{member_id}", response_model=MemberOut)
def patch_member(
    member_id: int,
    req: MemberPatch,
    db: Session = Depends(get_db),
    claims: dict = Depends(require_auth),
) -> models.Member:
    member = db.query(models.Member).get(member_id)
    if not member:
        raise HTTPException(status_code=404, detail="member not found")

    changes: list[str] = []
    for field in ("name", "email", "location", "tshirt_size", "address"):
        new_val = getattr(req, field)
        if new_val is None:
            continue
        new_val = new_val.strip()
        normalised = new_val if field == "name" else (new_val or None)
        if field == "name" and not normalised:
            raise HTTPException(status_code=400, detail="name cannot be blank")
        old_val = getattr(member, field)
        if normalised != old_val:
            changes.append(f"{field}: {old_val!r} -> {normalised!r}")
            setattr(member, field, normalised)

    if not changes:
        return member

    db.commit()
    db.refresh(member)
    _audit_team_edit(
        db, team_id=member.team_id,
        summary=f"Updated member #{member.id} ({member.name}): " + "; ".join(changes),
        reason=req.edit_reason,
        editor_email=_signed_in_email(claims),
    )
    db.commit()
    return member


def _find_raw_value(raw: dict, *needles: str) -> str | None:
    """Backfill helper — find the first JSON value whose key contains ALL
    given substrings (case-insensitive). Mirrors importer._find_col so the
    same matching logic applies post-hoc to the stored `raw` blob.
    """
    import re as _re
    def _norm(s: str) -> str:
        return _re.sub(r"\s+", " ", str(s or "").strip().lower())
    for k, v in raw.items():
        nk = _norm(k)
        if all(_norm(n) in nk for n in needles):
            if v is None:
                return None
            s = str(v).strip()
            return s if s else None
    return None


def _raw_address_slots(raw: dict) -> dict[int, str | None]:
    """Return mailing address values from a raw row keyed by slot index.

    MS Forms column naming:
      "Enter your mailing address if you opted for US or PH as location"  → slot 0 (mentor)
      "Enter your mailing address if you opted for US or PH as location1" → slot 1 (member 1)
      "Enter your mailing address if you opted for US or PH as location2" → slot 2 (member 2)
      ...
      "Enter your mailing address if you opted for US or PH as location5" → slot 5 (member 5)

    No trailing digit → slot 0 (mentor).  Trailing digit N → slot N (member N).
    """
    import re as _re

    def _norm(s: str) -> str:
        return _re.sub(r"\s+", " ", str(s or "").strip().lower())

    slots: dict[int, str | None] = {}
    for k, v in raw.items():
        nk = _norm(k)
        if ("mailing" in nk and "address" in nk) or "opted for us or ph" in nk:
            m = _re.search(r"(\d+)\s*$", k.strip())
            slot = int(m.group(1)) if m else 0  # no digit = mentor slot 0
            val = str(v).strip() if v is not None else None
            slots[slot] = val if val else None
    return slots


@app.post("/api/admin/backfill-mentor-locations")
def backfill_mentor_locations(
    db: Session = Depends(get_db),
    claims: dict = Depends(require_auth),
) -> dict:
    """One-shot: read each Team.raw and populate mentor_location /
    mentor_tshirt_size / mentor_address / member.address for records
    where those columns are blank.

    Useful after the importer fix lands: older registrations imported
    before these fields were captured still have NULL in the columns,
    but the original Excel row sits in the `raw` JSON blob — so we can
    recover the values without a re-upload (which would wipe organizer
    edits made on the dashboard).
    """
    teams = db.query(models.Team).all()
    locations_set = 0
    tshirts_set = 0
    mentor_addresses_set = 0
    member_locations_set = 0
    member_addresses_set = 0
    examined = 0
    for t in teams:
        if not isinstance(t.raw, dict) or not t.raw:
            continue
        examined += 1
        if not t.mentor_location:
            loc = _find_raw_value(t.raw, "mentor", "location")
            if loc:
                t.mentor_location = loc
                locations_set += 1
        if not t.mentor_tshirt_size:
            ts = _find_raw_value(t.raw, "mentor", "shirt")
            if ts:
                t.mentor_tshirt_size = ts
                tshirts_set += 1
        addr_slots = _raw_address_slots(t.raw)
        if not t.mentor_address:
            addr = addr_slots.get(0)
            if addr:
                t.mentor_address = addr
                mentor_addresses_set += 1
        for m in t.members:
            pos = m.position
            if not m.location:
                loc = (
                    _find_raw_value(t.raw, f"member {pos}", "location")
                    or _find_raw_value(t.raw, f"member{pos}", "location")
                )
                if loc:
                    m.location = loc
                    member_locations_set += 1
            if not m.address:
                addr = addr_slots.get(pos + 1)  # form slots are 1-indexed; member.position is 0-indexed
                if addr:
                    m.address = addr
                    member_addresses_set += 1
    db.commit()
    return {
        "teams_total": len(teams),
        "teams_with_raw": examined,
        "mentor_locations_set": locations_set,
        "mentor_tshirt_sizes_set": tshirts_set,
        "mentor_addresses_set": mentor_addresses_set,
        "member_locations_set": member_locations_set,
        "member_addresses_set": member_addresses_set,
        "performed_by": _signed_in_email(claims),
    }


@app.delete("/api/members/{member_id}")
def delete_member(
    member_id: int,
    db: Session = Depends(get_db),
    claims: dict = Depends(require_auth),
) -> dict:
    member = db.query(models.Member).get(member_id)
    if not member:
        raise HTTPException(status_code=404, detail="member not found")
    team_id = member.team_id
    summary = f"Removed member #{member.id} ({member.name} <{member.email or 'no-email'}>)"
    db.delete(member)
    db.commit()
    _audit_team_edit(
        db, team_id=team_id, summary=summary, reason=None,
        editor_email=_signed_in_email(claims),
    )
    db.commit()
    return {"deleted": True, "member_id": member_id}
